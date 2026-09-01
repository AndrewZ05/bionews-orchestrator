#!/usr/bin/env python3
"""
Backfill + incremental load of email-delivered Ad-Ops reports into BigQuery.

Generalises the proven CM360 loader (scripts/cms_xlsx_backfill.py) to multiple
sources that land in gs://email-file-ingestion/<source>/<YYYY-MM-DD>/*. Each
source is described by a SOURCE config below; the convert -> per-file Parquet ->
GCS -> bq load + dedup pipeline is shared.

Supported sources:
  dv   DoubleVerify   CSV  -> csm_data.dv_report
  ttd  The Trade Desk  XLSX -> csm_data.ttd_ad_group_performance

Design (same as cms):
  - RAW wide-union columns (snake_cased source headers), all variants merged.
  - One Parquet per source file at
    gs://{GCS_BUCKET}/csm/parquet/{source}/{stem}.parquet -> the per-file Parquet
    existence is the INCREMENTAL marker (re-run only processes new files).
  - Explicit pyarrow schema on write so every Parquet has identical physical
    types (avoids the BQ "type INT64 != DOUBLE" load failures seen with cms).
  - Dedup: CREATE OR REPLACE ... QUALIFY ROW_NUMBER() OVER (PARTITION BY natural
    key ORDER BY source_filename DESC) -- latest delivered file wins.
  - DoubleVerify only: early files (2025-02..2025-06-13) are an HTML login page,
    not data. Detected and SKIPPED (logged), never loaded.

Datasets reused: csm_staging / csm_data (per user decision), table names
prefixed per source.

Modes (same flags as cms loader): --dry-run | --convert-only | --confirm-write,
plus --reload-all. --source dv|ttd selects the source.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import tempfile
from collections import Counter
from concurrent.futures import ProcessPoolExecutor, as_completed

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq_write

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    print("python-dotenv not installed; run inside the project venv", file=sys.stderr)
    raise

import openpyxl
from google.cloud import bigquery, storage

SRC_BUCKET = "email-file-ingestion"
PARQUET_PREFIX = "csm/parquet/"  # under GCS_BUCKET, namespaced by source below


# --------------------------------------------------------------------------- #
# Per-source configuration
# --------------------------------------------------------------------------- #
SOURCES = {
    "dv": {
        "prefix": "doubleverify/",
        "ext": ".csv",
        "reader": "csv",
        "table": "dv_report",
        # natural key for dedup: DV reports are daily per placement/advertiser.
        # Different variants carry different dims; key on what's broadly present.
        "dedup_keys": [
            "date",
            "advertiser_name",
            "campaign_name",
            "placement_name",
            "media_property",
            "media_type",
            "device_delivery_type",
        ],
        "partition_col": "date",
        # int vs float vs date typing for snake_cased columns
        "int_cols": {
            "monitored_ads",
            "measured_impressions",
            "viewable_impressions",
            "blocks",
            "requests",
            "fraud_sivt_incidents",
            "fraud_sivt_blocks",
            "out_of_geo_blocks",
            "out_of_geo_incidents",
            "brand_suitability_incidents",
            "brand_suitability_blocks",
            "unique_incidents",
            "authentic_ads",
            "evaluations",
            "filters",
            "brand_suitability_filters",
            "fraud_sivt_filters",
            "out_of_geo_filters",
            "eligible_impressions",
            "keyword_blocks",
            "keyword_incidents",
            "brand_risk_floor_incidents",
            "pmx_authentic_passed_impressions",
            "total_net_pmx_measured_impressions",
        },
        "float_cols": {
            "measurement_rate",
            "viewable_rate",
            "authentic_rate",
            "block_rate",
            "filter_rate",
            "brand_suitability_incident_rate",
            "fraud_sivt_incident_rate",
            "out_of_geo_incident_rate",
            "brand_suitability_block_rate",
            "fraud_sivt_block_rate",
            "out_of_geo_block_rate",
            "brand_suitability_filter_rate",
            "fraud_sivt_filter_rate",
            "out_of_geo_filter_rate",
        },
        "date_cols": {"date"},
        "skip_html": True,  # DV early feed is broken HTML
    },
    "ttd": {
        "prefix": "tradedesk/",
        "ext": ".xlsx",
        "reader": "xlsx_data_sheet",  # sheet ending in '_data'
        "table": "ttd_ad_group_performance",
        "dedup_keys": [
            "date",
            "campaign",
            "ad_group",
            "creative",
            "media_type",
            "ad_format",
            "inventory_contract",
        ],
        "partition_col": "date",
        "int_cols": {
            "impressions",
            "clicks",
            "bids",
            "ad_plays",
            "player_starts",
            "player_completed_views",
            "sampled_tracked_impressions",
            "sampled_viewed_impressions",
            "quality_reach_index_measured_impressions",
            "normalized_relevance_measured_impression_count_filtered_hidden",
        },
        "float_cols": {
            "advertiser_cost_adv_currency",
            "partner_cost_adv_currency",
            "total_bid_amount_adv_currency",
            "normalized_relevance_score_filtered_hidden",
            "normalized_relevance_score_v2_filtered",
            "quality_reach_index_raw",
        },
        "date_cols": {"date"},
        "skip_html": False,
    },
}

# Derived/lineage columns added to every row (not literal source headers).
DERIVED_COLS = ["source_filename", "source_gcs_path", "report_date_partition"]


def safe_stem(stem: str) -> str:
    """Filesystem/object-safe, UNIQUE name for a source file stem.

    TTD filenames are long and only differ in a trailing date/id suffix, so a
    naive truncation collides (all files of a campaign map to one object and
    clobber each other). Keep a 100-char prefix AND append a short hash of the
    FULL stem to guarantee uniqueness.
    """
    import hashlib

    cleaned = re.sub(r"[^\w.-]+", "_", stem)
    h = hashlib.sha1(stem.encode("utf-8")).hexdigest()[:10]
    return f"{cleaned[:100]}__{h}"


# --------------------------------------------------------------------------- #
def snake(name: str) -> str:
    s = name.strip().lower()
    s = s.replace("%", " pct ").replace("/", " ").replace("(", " ").replace(")", " ")
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def is_html(raw: bytes) -> bool:
    head = raw[:300].decode("utf-8-sig", "replace").lstrip().lower()
    return head.startswith("<!doctype") or "<html" in head[:200]


def _arrow_type(c: str, cfg: dict):
    if c in cfg["int_cols"]:
        return pa.int64()
    if c in cfg["float_cols"]:
        return pa.float64()
    if c in cfg["date_cols"]:
        return pa.date32()
    return pa.string()


def _arrow_schema(cols, cfg):
    return pa.schema([pa.field(c, _arrow_type(c, cfg)) for c in cols])


def read_source_file(local_path: str, reader: str) -> pd.DataFrame | None:
    if reader == "csv":
        return pd.read_csv(
            local_path, dtype=object, encoding="utf-8-sig", on_bad_lines="skip"
        )
    if reader == "xlsx_data_sheet":
        wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
        data_sheets = [s for s in wb.sheetnames if s.endswith("_data")]
        sheet = data_sheets[0] if data_sheets else wb.sheetnames[0]
        wb.close()
        return pd.read_excel(
            local_path, sheet_name=sheet, dtype=object, engine="openpyxl"
        )
    raise ValueError(f"unknown reader {reader}")


# Worker: convert ONE source file to a local Parquet (subprocess-safe).
def convert_one(task: dict) -> dict:
    src = task["src_blob"]
    local = task["local"]
    full_cols = task["full_cols"]
    out_dir = task["out_local_dir"]
    cfg = task["cfg"]
    source = task["source"]
    try:
        with open(local, "rb") as f:
            head = f.read(300)
        if cfg["skip_html"] and is_html(head):
            return {"src": src, "status": "html_skipped", "rows": 0}
        df = read_source_file(local, cfg["reader"])
        if df is None or df.empty:
            return {"src": src, "status": "empty", "rows": 0}
        df.columns = [snake(str(c)) for c in df.columns]
        df = df.loc[:, [c for c in df.columns if c and not c.startswith("unnamed")]]
        # drop fully-blank trailing rows
        if len(df.columns):
            first = df.iloc[:, 0].astype(str)
            df = df[
                first.notna() & (first.str.strip() != "") & (first.str.lower() != "nan")
            ]

        df = df.reindex(columns=full_cols)
        for c in df.columns:
            if c in cfg["int_cols"]:
                df[c] = pd.to_numeric(df[c], errors="coerce").round().astype("Int64")
            elif c in cfg["float_cols"]:
                df[c] = pd.to_numeric(df[c], errors="coerce").astype("float64")
            elif c in cfg["date_cols"]:
                df[c] = pd.to_datetime(df[c], errors="coerce").dt.date
            else:
                df[c] = df[c].astype("string")

        stem = os.path.splitext(os.path.basename(src))[0]
        m = re.search(r"/(\d{4}-\d{2}-\d{2})/", src)
        df["source_filename"] = os.path.basename(src)
        df["source_gcs_path"] = f"gs://{SRC_BUCKET}/{src}"
        df["report_date_partition"] = m.group(1) if m else None

        os.makedirs(out_dir, exist_ok=True)
        safe = safe_stem(stem)  # unique even when names share a long prefix
        local_pq = os.path.join(out_dir, f"{source}__{safe}.parquet")
        table = pa.Table.from_pandas(
            df, schema=_arrow_schema(df.columns, cfg), preserve_index=False
        )
        pq_write.write_table(table, local_pq)
        return {
            "src": src,
            "status": "ok",
            "rows": len(df),
            "local_pq": local_pq,
            "object": f"{PARQUET_PREFIX}{source}/{safe}.parquet",
        }
    except Exception as e:  # noqa: BLE001
        return {"src": src, "status": "error", "rows": 0, "error": repr(e)[:200]}


def _header_of(local_path: str, cfg: dict) -> list[str]:
    """Read just the header row of a local file (CSV or xlsx _data sheet)."""
    with open(local_path, "rb") as f:
        head = f.read(300)
    if cfg["skip_html"] and is_html(head):
        return []
    if cfg["reader"] == "csv":
        import csv as _csv

        with open(local_path, encoding="utf-8-sig", errors="replace") as f:
            try:
                return next(_csv.reader(f))
            except StopIteration:
                return []
    wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
    ds = [s for s in wb.sheetnames if s.endswith("_data")] or wb.sheetnames
    ws = wb[ds[0]]
    hdr: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        hdr = [str(c).strip() for c in row if c is not None]
        break
    wb.close()
    return hdr


def build_union_local(name_path_pairs, cfg) -> list[str]:
    """Wide-union snake_cased columns from the FULL set of local files."""
    union, seen = [], set()
    for _name, local_path in name_path_pairs:
        for c in _header_of(local_path, cfg):
            sc = snake(str(c))
            if sc and sc not in seen:
                seen.add(sc)
                union.append(sc)
    return union


def list_blobs(client, prefix, ext, limit=0):
    blobs = [
        b
        for b in client.list_blobs(SRC_BUCKET, prefix=prefix)
        if b.name.lower().endswith(ext)
    ]
    blobs.sort(key=lambda b: b.name)
    return blobs[:limit] if limit else blobs


def existing_stems(client, gcs_bucket, source):
    out = set()
    for b in client.list_blobs(gcs_bucket, prefix=f"{PARQUET_PREFIX}{source}/"):
        if b.name.endswith(".parquet"):
            out.add(os.path.splitext(os.path.basename(b.name))[0])
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, choices=list(SOURCES))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--convert-only", action="store_true")
    ap.add_argument("--confirm-write", action="store_true")
    ap.add_argument("--reload-all", action="store_true")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--workers", type=int, default=0)
    ap.add_argument(
        "--out", default=os.path.join(tempfile.gettempdir(), "adops_backfill")
    )
    args = ap.parse_args()

    if not (args.dry_run or args.convert_only or args.confirm_write):
        print(
            "Pick a mode: --dry-run | --convert-only | --confirm-write", file=sys.stderr
        )
        return 2

    cfg = SOURCES[args.source]
    source = args.source
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    project = os.getenv("GCP_PROJECT_ID", "bi-data-391216")
    gcs_bucket = os.getenv("GCS_BUCKET")
    out = os.path.join(args.out, source)
    dl_dir, pq_dir = os.path.join(out, "dl"), os.path.join(out, "pq")
    os.makedirs(dl_dir, exist_ok=True)
    os.makedirs(pq_dir, exist_ok=True)

    sc = storage.Client(project=project)
    all_blobs = list_blobs(sc, cfg["prefix"], cfg["ext"], args.limit)
    print(f"[{source}] source files: {len(all_blobs)}")

    # --- download every (limited) source file ONCE, locally ---
    # Build the wide-union from the FULL corpus (a new file might introduce a
    # column), then convert. Downloading once avoids the double-fetch.
    def local_for(b):
        return os.path.join(dl_dir, re.sub(r"[^\w.-]+", "_", os.path.basename(b.name)))

    for n, b in enumerate(all_blobs, 1):
        lp = local_for(b)
        if not os.path.exists(lp):
            b.download_to_filename(lp)
        if n % 300 == 0:
            print(f"  downloaded {n}/{len(all_blobs)}")

    # wide-union columns from local files (header-only read)
    union_cols = build_union_local([(b.name, local_for(b)) for b in all_blobs], cfg)
    full_cols = union_cols + [c for c in DERIVED_COLS if c not in union_cols]
    print(f"[{source}] wide-union columns: {len(full_cols)}")

    # incremental skip (after union so the union always reflects the full corpus)
    blobs = all_blobs
    if (args.confirm_write or args.convert_only) and not args.reload_all and gcs_bucket:
        done = existing_stems(sc, gcs_bucket, source)
        before = len(blobs)
        # The GCS object basename is exactly safe_stem(stem) (the source folder
        # already namespaces it), so compare against that -- NOT a "source__"
        # prefixed form, which never matched and broke incremental detection.
        blobs = [
            b
            for b in blobs
            if safe_stem(os.path.splitext(os.path.basename(b.name))[0]) not in done
        ]
        print(
            f"[{source}] incremental: {len(done)} done, {before-len(blobs)} skipped, "
            f"{len(blobs)} to process"
        )

    tasks = [
        {
            "src_blob": b.name,
            "local": local_for(b),
            "full_cols": full_cols,
            "out_local_dir": pq_dir,
            "cfg": cfg,
            "source": source,
        }
        for b in blobs
    ]

    workers = args.workers or min(6, max(1, (os.cpu_count() or 4) - 1))
    print(f"[{source}] converting {len(tasks)} file(s) with {workers} worker(s)...")
    results = []
    if tasks:
        with ProcessPoolExecutor(max_workers=workers) as ex:
            futs = [ex.submit(convert_one, t) for t in tasks]
            for i, fut in enumerate(as_completed(futs), 1):
                results.append(fut.result())
                if i % 200 == 0:
                    print(f"  converted {i}/{len(tasks)}")

    by_status = Counter(r["status"] for r in results)
    rows = sum(r["rows"] for r in results if r["status"] == "ok")
    print(f"\n[{source}] convert: {dict(by_status)}  total rows: {rows:,}")
    for r in [x for x in results if x["status"] in ("error",)][:15]:
        print(f"  [error] {r['src']} :: {r.get('error')}")
    if by_status.get("html_skipped"):
        print(
            f"  html_skipped: {by_status['html_skipped']} broken DV files (logged, not loaded)"
        )

    if args.dry_run:
        print(f"[{source}][dry-run] local Parquet under {pq_dir}; nothing uploaded.")
        return 0

    if not gcs_bucket:
        print("ERROR: GCS_BUCKET not set", file=sys.stderr)
        return 2
    bucket = sc.bucket(gcs_bucket)
    up = 0
    for r in results:
        if r["status"] != "ok":
            continue
        bucket.blob(r["object"]).upload_from_filename(r["local_pq"])
        up += 1
        if up % 200 == 0:
            print(f"  uploaded {up}")
    print(f"[{source}] uploaded {up} Parquet object(s)")

    if args.convert_only:
        print(f"[{source}][convert-only] BQ load skipped.")
        return 0

    bq = bigquery.Client(project=project)
    _load(bq, project, gcs_bucket, source, cfg, full_cols)
    print(f"[{source}] DONE.")
    return 0


def _ensure_dataset(bq, name):
    from google.cloud.exceptions import NotFound

    ds_id = f"{bq.project}.{name}"
    try:
        bq.get_dataset(ds_id)
    except NotFound:
        ds = bigquery.Dataset(ds_id)
        ds.location = "US"
        bq.create_dataset(ds, timeout=30)
        print(f"created dataset {ds_id}")


def _bq_schema(full_cols, cfg):
    fields = []
    for c in full_cols:
        if c in cfg["int_cols"]:
            t = "INT64"
        elif c in cfg["float_cols"]:
            t = "FLOAT64"
        elif c in cfg["date_cols"]:
            t = "DATE"
        else:
            t = "STRING"
        fields.append(bigquery.SchemaField(c, t))
    return fields


def _load(bq, project, gcs_bucket, source, cfg, full_cols):
    _ensure_dataset(bq, "csm_staging")
    _ensure_dataset(bq, "csm_data")
    table = cfg["table"]
    stg = f"{project}.csm_staging.{table}__stg"
    prod = f"{project}.csm_data.{table}"
    uri = f"gs://{gcs_bucket}/{PARQUET_PREFIX}{source}/*.parquet"

    job = bq.load_table_from_uri(
        uri,
        stg,
        job_config=bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.PARQUET,
            write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
            schema=_bq_schema(full_cols, cfg),
        ),
    )
    job.result()
    n_stg = bq.get_table(stg).num_rows
    print(f"[{source}] staged {n_stg:,} -> {stg}")

    # dedup keys: only those that actually exist in the union
    keys = [k for k in cfg["dedup_keys"] if k in full_cols] or ["source_filename"]
    part = ", ".join(keys)
    pcol = cfg["partition_col"]
    part_clause = f"PARTITION BY {pcol}" if pcol in full_cols else ""
    sql = f"""
    CREATE OR REPLACE TABLE `{prod}`
    {part_clause}
    AS
    SELECT * EXCEPT(_rn) FROM (
      SELECT *, ROW_NUMBER() OVER (PARTITION BY {part} ORDER BY source_filename DESC) AS _rn
      FROM `{stg}`
    ) WHERE _rn = 1
    """
    bq.query(sql).result()
    n_prod = bq.get_table(prod).num_rows
    print(
        f"[{source}] de-duped -> {prod}: {n_prod:,} rows ({n_stg-n_prod:,} collapsed)"
    )


if __name__ == "__main__":
    raise SystemExit(main())
