#!/usr/bin/env python3
"""
Backfill + incremental load of the email-delivered CM360 .xlsx corpus into BigQuery.

Source: gs://email-file-ingestion/cm360/<YYYY-MM-DD>/<accountid>_*.xlsx

Pipeline:
  1. CONVERT  each .xlsx -> one snake_cased, wide-union Parquet, written to
     gs://{GCS_BUCKET}/csm/parquet/{grain}/{account}/{stem}.parquet
     Parallelised across files (openpyxl is slow single-threaded).
     The per-source-file Parquet is the unit of INCREMENTAL detection: a source
     file is "already done" iff its destination Parquet object exists.
  3 grains, 3 destination tables:
     daily   -> csm_data.cms_campaign_performance        (partition: date)
     monthly -> csm_data.cms_campaign_performance_monthly (string month)
     period  -> csm_data.cms_campaign_performance_period  (placement totals)
  2. LOAD     bq load all Parquet for a grain into csm_staging.<t>__stg, then
     CREATE OR REPLACE the production table de-duplicated on the natural key
     (latest source file wins -- attribution is revised retroactively).

Columns are RAW (snake_cased source headers), union/superset across all header
variants, plus derived cols: site (=COALESCE(site_cm360,site_dcm)), grain,
source_account_id, source_filename, report_window, source_gcs_path,
ingest_loaded_at.

Modes:
  --dry-run                Convert a sample to LOCAL Parquet only. No GCS/BQ writes.
  --convert-only           Convert (parallel) + upload Parquet to GCS. No BQ.
  --confirm-write          Full run: convert (incremental) + upload + BQ load+dedup.
  --reload-all             With --confirm-write: ignore incremental skip, reconvert all.

Usage:
  python scripts/cms_xlsx_backfill.py --dry-run --limit 40 --manifest <m.json>
  python scripts/cms_xlsx_backfill.py --confirm-write --manifest <m.json>
  python scripts/cms_xlsx_backfill.py --confirm-write           # nightly incremental
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from collections import Counter
from concurrent.futures import ProcessPoolExecutor, as_completed

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq_write

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    print("python-dotenv not installed; run inside the project venv", file=sys.stderr)
    raise

import openpyxl
from google.cloud import bigquery, storage

SRC_BUCKET = "email-file-ingestion"
SRC_PREFIX = "cm360/"
PARQUET_PREFIX = "csm/parquet/"  # under GCS_BUCKET
SCAN_ROWS = 30
ANCHOR_TOKENS = ("impression", "placement", "campaign")

PATH_RE = re.compile(r"cm360/(\d{4}-\d{2}-\d{2})/(\d+)_(.+?)\.xlsx$", re.IGNORECASE)
WINDOW_RE = re.compile(r"(MTD|QTD|YTD|Daily|Weekly|Monthly)", re.IGNORECASE)

DERIVED_COLS = [
    "site",
    "grain",
    "source_account_id",
    "source_filename",
    "report_window",
    "source_gcs_path",
    "ingest_loaded_at",
]

GRAIN_TABLE = {
    "daily": "cms_campaign_performance",
    "monthly": "cms_campaign_performance_monthly",
    "period": "cms_campaign_performance_period",
}

# Natural key for de-dup within a grain. The same logical row recurs across many
# overlapping report files (a YTD report re-emits every prior day weekly); keep
# the row from the most recently generated file (metrics get revised
# retroactively). The key MUST include every dimension that can vary within a
# single file -- creative/ad/platform/country -- or distinct breakdown rows get
# silently collapsed (verified: coarse key dropped ~10k legit rows). site is the
# coalesced column. Columns absent from a given file are NULL and group together.
_DIMS = [
    "campaign_id",
    "placement_id",
    "creative_id",
    "ad_id",
    "site",
    "platform_type",
    "country",
]
DEDUP_KEYS = {
    "daily": ["source_account_id", "date"] + _DIMS,
    "monthly": ["source_account_id", "month"] + _DIMS,
    "period": ["source_account_id"] + _DIMS,
}

# integer/float columns to coerce on load (everything else stays STRING)
INT_COLS = {
    "impressions",
    "clicks",
    "campaign_id",
    "placement_id",
    "creative_id",
    "ad_id",
    "package_roadblock_id",
    "package_roadblock_total_booked_units",
    "placement_total_booked_units",
    "advertiser_id",
    "site_id_cm360",
    "active_view_viewable_impressions",
    "active_view_measurable_impressions",
    "active_view_eligible_impressions",
    "active_view_pct_viewable_impressions",
    "active_view_audible_measurable_impressions",
}
FLOAT_COLS = {
    "click_rate",
    "total_conversions",
    "average_time",
    "total_display_time",
    "total_interaction_time",
}


def _arrow_schema(cols) -> "pa.Schema":
    """Explicit pyarrow schema -> identical physical types across every Parquet,
    even for all-null columns (date->date32, ints->int64, floats->float64)."""
    fields = []
    for c in cols:
        if c in INT_COLS:
            t = pa.int64()
        elif c in FLOAT_COLS:
            t = pa.float64()
        elif c == "date":
            t = pa.date32()
        else:
            t = pa.string()
        fields.append(pa.field(c, t))
    return pa.schema(fields)


# --------------------------------------------------------------------------- #
# Parsing helpers (verified in prior dry-runs)
# --------------------------------------------------------------------------- #
def snake(name: str) -> str:
    # Preserve a leading '%' as 'pct' so the Active View viewable-% rate column
    # does not collide with the viewable-impressions count column.
    s = name.strip().lower()
    s = s.replace("(cm360)", "cm360").replace("(dcm)", "dcm")
    s = s.replace("%", " pct ")
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def find_header_idx(rows: list[tuple]) -> tuple[int | None, str | None]:
    """Return (header_row_index, grain): daily|monthly|period."""
    for i, row in enumerate(rows):
        cells = ["" if c is None else str(c).strip() for c in row]
        nonempty = [c for c in cells if c]
        if not nonempty:
            continue
        joined = " | ".join(nonempty).lower()
        is_anchor = any(tok in joined for tok in ANCHOR_TOKENS)
        first = cells[0].lower()
        if not is_anchor or "metric" in first or "field" in first:
            continue
        if "Date" in cells:
            return i, "daily"
        if "Month" in cells:
            return i, "monthly"
        return i, "period"
    return None, None


def read_file_to_df(local_path: str, gcs_name: str) -> tuple[pd.DataFrame | None, str]:
    """Read one .xlsx into a snake_cased DataFrame + derived cols. Returns (df, grain)."""
    wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    head = [
        r
        for _, r in zip(
            range(SCAN_ROWS),
            ws.iter_rows(min_row=1, max_row=SCAN_ROWS, values_only=True),
        )
    ]
    wb.close()
    hidx, grain = find_header_idx(head)
    if hidx is None:
        return None, "noheader"

    df = pd.read_excel(
        local_path, sheet_name=0, skiprows=hidx, dtype=object, engine="openpyxl"
    )
    df.columns = [snake(str(c)) for c in df.columns]
    df = df.loc[:, [c for c in df.columns if c and not c.startswith("unnamed")]]

    if len(df.columns):
        first = df.iloc[:, 0].astype(str)
        # single combined mask -> avoids the reindex warning from chaining two
        # filters where the second mask was built from the pre-filter index
        keep = (
            first.notna()
            & (first.str.strip() != "")
            & ~first.str.contains("grand total|total", case=False, na=False)
        )
        df = df[keep]

    time_col = {"daily": "date", "monthly": "month"}.get(grain)
    if time_col and time_col in df.columns:
        df = df[df[time_col].notna()]

    # coalesced site (site_cm360 modern / site_dcm legacy; never co-occur)
    sc = df["site_cm360"] if "site_cm360" in df.columns else None
    sd = df["site_dcm"] if "site_dcm" in df.columns else None
    if sc is not None and sd is not None:
        df["site"] = sc.fillna(sd)
    elif sc is not None:
        df["site"] = sc
    elif sd is not None:
        df["site"] = sd

    m = PATH_RE.search(gcs_name)
    win = WINDOW_RE.search(os.path.basename(gcs_name))
    df["grain"] = grain
    df["source_account_id"] = m.group(2) if m else None
    df["source_filename"] = os.path.basename(gcs_name)
    df["report_window"] = win.group(1).upper() if win else None
    df["source_gcs_path"] = f"gs://{SRC_BUCKET}/{gcs_name}"
    df["ingest_loaded_at"] = None  # stamped at BQ load
    return df, grain


def build_union_from_manifest(manifest_path: str) -> list[str]:
    with open(manifest_path, encoding="utf-8") as f:
        man = json.load(f)
    cols, seen, out = [snake(n) for n, _ in man["union_raw"]], set(), []
    for c in cols:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out


def parquet_object_name(grain: str, account: str, stem: str) -> str:
    return f"{PARQUET_PREFIX}{grain}/{account}/{stem}.parquet"


# --------------------------------------------------------------------------- #
# Worker: convert ONE source file to a local Parquet (run in a subprocess).
# Pure function of its inputs so it parallelises cleanly.
# --------------------------------------------------------------------------- #
def convert_one(task: dict) -> dict:
    """task: {src_blob, local_xlsx, full_cols, out_local_dir}. Returns a result dict."""
    src = task["src_blob"]
    local_xlsx = task["local_xlsx"]
    full_cols = task["full_cols"]
    out_dir = task["out_local_dir"]
    try:
        df, grain = read_file_to_df(local_xlsx, src)
        if grain == "noheader" or df is None:
            return {"src": src, "status": "noheader", "grain": grain, "rows": 0}
        if df.empty:
            return {"src": src, "status": "empty", "grain": grain, "rows": 0}
        df = df.reindex(columns=full_cols)
        # type coercion (best-effort; bad cells -> NaN). For INT_COLS, go via a
        # nullable float and round before Int64 -- some "int" cells arrive as
        # floats ('17.0') or junk, and a direct .astype('Int64') raises
        # "cannot safely cast non-equivalent ... to int64".
        # FLOAT_COLS/INT_COLS must get a STABLE physical type in every Parquet,
        # else BQ rejects the load: a file whose total_conversions is all-null or
        # all-whole writes INT64, another writes DOUBLE -> "type INT64 does not
        # match target cpp_type DOUBLE". Force float64 / Int64 explicitly.
        for c in df.columns:
            if c in INT_COLS:
                df[c] = pd.to_numeric(df[c], errors="coerce").round().astype("Int64")
            elif c in FLOAT_COLS:
                df[c] = pd.to_numeric(df[c], errors="coerce").astype("float64")
            elif c == "date":
                df[c] = pd.to_datetime(df[c], errors="coerce").dt.date
            else:
                df[c] = df[c].astype("string")
        m = PATH_RE.search(src)
        account = m.group(2) if m else "unknown"
        stem = os.path.splitext(os.path.basename(src))[0]
        os.makedirs(out_dir, exist_ok=True)
        local_pq = os.path.join(out_dir, f"{grain}__{account}__{stem}.parquet")
        # Write with an EXPLICIT pyarrow schema so every Parquet has identical
        # physical column types regardless of content. Without this, an all-null
        # column's arrow type varies per file (date -> timestamp vs date32,
        # total_conversions -> int64 vs double) and BQ rejects the mixed load.
        table = pa.Table.from_pandas(
            df, schema=_arrow_schema(df.columns), preserve_index=False
        )
        pq_write.write_table(table, local_pq)
        return {
            "src": src,
            "status": "ok",
            "grain": grain,
            "rows": len(df),
            "local_pq": local_pq,
            "object": parquet_object_name(grain, account, stem),
        }
    except Exception as e:  # noqa: BLE001
        return {
            "src": src,
            "status": "error",
            "grain": "?",
            "rows": 0,
            "error": repr(e)[:200],
        }


# --------------------------------------------------------------------------- #
def list_source_blobs(client: storage.Client, limit: int = 0) -> list:
    blobs = [
        b
        for b in client.list_blobs(SRC_BUCKET, prefix=SRC_PREFIX)
        if b.name.lower().endswith(".xlsx")
    ]
    blobs.sort(key=lambda b: b.name)
    return blobs[:limit] if limit else blobs


def existing_parquet_stems(client: storage.Client, gcs_bucket: str) -> set[str]:
    """Set of source stems already converted (object path -> stem)."""
    out = set()
    for b in client.list_blobs(gcs_bucket, prefix=PARQUET_PREFIX):
        if b.name.endswith(".parquet"):
            out.add(os.path.splitext(os.path.basename(b.name))[0])
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--dry-run", action="store_true", help="local Parquet only; no GCS/BQ"
    )
    ap.add_argument(
        "--convert-only", action="store_true", help="convert + upload Parquet; no BQ"
    )
    ap.add_argument(
        "--confirm-write", action="store_true", help="full run incl. BQ load"
    )
    ap.add_argument("--reload-all", action="store_true", help="ignore incremental skip")
    ap.add_argument(
        "--manifest", default="", help="header_manifest.json (wide-union cols)"
    )
    ap.add_argument("--limit", type=int, default=0, help="first N source files (0=all)")
    ap.add_argument(
        "--workers", type=int, default=0, help="parallel convert workers (0=auto)"
    )
    ap.add_argument(
        "--out", default=os.path.join(tempfile.gettempdir(), "cms_backfill")
    )
    args = ap.parse_args()

    if not (args.dry_run or args.convert_only or args.confirm_write):
        print(
            "Pick a mode: --dry-run | --convert-only | --confirm-write", file=sys.stderr
        )
        return 2

    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    project = os.getenv("GCP_PROJECT_ID", "bi-data-391216")
    gcs_bucket = os.getenv("GCS_BUCKET")
    os.makedirs(args.out, exist_ok=True)
    dl_dir = os.path.join(args.out, "dl")
    pq_dir = os.path.join(args.out, "pq")
    os.makedirs(dl_dir, exist_ok=True)
    os.makedirs(pq_dir, exist_ok=True)

    storage_client = storage.Client(project=project)

    if not args.manifest or not os.path.exists(args.manifest):
        print(
            "ERROR: --manifest header_manifest.json is required (defines wide-union columns)",
            file=sys.stderr,
        )
        return 2
    union_cols = build_union_from_manifest(args.manifest)
    full_cols = union_cols + [c for c in DERIVED_COLS if c not in union_cols]
    print(f"wide-union columns: {len(full_cols)}")

    blobs = list_source_blobs(storage_client, args.limit)
    print(f"source .xlsx files: {len(blobs)}")

    # incremental: skip files whose Parquet already exists (unless --reload-all/--dry-run)
    skip_done = (args.confirm_write or args.convert_only) and not args.reload_all
    done_stems: set[str] = set()
    if skip_done and gcs_bucket:
        done_stems = existing_parquet_stems(storage_client, gcs_bucket)
        before = len(blobs)
        blobs = [
            b
            for b in blobs
            if os.path.splitext(os.path.basename(b.name))[0] not in done_stems
        ]
        print(
            f"incremental: {len(done_stems)} already converted; "
            f"{before - len(blobs)} skipped, {len(blobs)} to process"
        )

    if not blobs:
        # No new source files. In --confirm-write we still fall through to the
        # BQ load step so any previously-uploaded Parquet is (re)loaded/deduped.
        print("nothing new to convert.")

    # download (sequential -- network bound) then convert (parallel -- CPU bound)
    tasks = []
    for n, b in enumerate(blobs, 1):
        local = os.path.join(dl_dir, os.path.basename(b.name))
        if not os.path.exists(local):
            b.download_to_filename(local)
        tasks.append(
            {
                "src_blob": b.name,
                "local_xlsx": local,
                "full_cols": full_cols,
                "out_local_dir": pq_dir,
            }
        )
        if n % 200 == 0:
            print(f"  downloaded {n}/{len(blobs)}")

    # openpyxl is RAM-hungry on large files; this is memory-bound, not CPU-bound.
    # Cap workers low to avoid OOM / BrokenProcessPool. Override with --workers.
    workers = args.workers or min(6, max(1, (os.cpu_count() or 4) - 1))
    print(f"converting {len(tasks)} file(s) with {workers} worker(s)...")
    results: list[dict] = []
    if tasks:
        with ProcessPoolExecutor(max_workers=workers) as ex:
            futs = [ex.submit(convert_one, t) for t in tasks]
            for i, fut in enumerate(as_completed(futs), 1):
                results.append(fut.result())
                if i % 200 == 0:
                    print(f"  converted {i}/{len(tasks)}")

    by_status = Counter(r["status"] for r in results)
    by_grain = Counter(r["grain"] for r in results if r["status"] == "ok")
    total_rows = sum(r["rows"] for r in results if r["status"] == "ok")
    print("\n" + "=" * 70)
    print(f"convert results: {dict(by_status)}")
    print(f"by grain (ok):   {dict(by_grain)}")
    print(f"total rows:      {total_rows:,}")
    for r in [x for x in results if x["status"] in ("error", "noheader")][:20]:
        print(f"  [{r['status']}] {r['src']} {r.get('error','')}")

    if args.dry_run:
        print(f"\n[dry-run] local Parquet under {pq_dir}; nothing uploaded.")
        return 0

    # ---- upload Parquet to GCS ----
    if not gcs_bucket:
        print("ERROR: GCS_BUCKET not set in .env", file=sys.stderr)
        return 2
    bucket = storage_client.bucket(gcs_bucket)
    uploaded = 0
    for r in results:
        if r["status"] != "ok":
            continue
        bucket.blob(r["object"]).upload_from_filename(r["local_pq"])
        uploaded += 1
        if uploaded % 200 == 0:
            print(f"  uploaded {uploaded}")
    print(
        f"uploaded {uploaded} Parquet object(s) to gs://{gcs_bucket}/{PARQUET_PREFIX}"
    )

    if args.convert_only:
        print("[convert-only] Parquet in GCS; BQ load skipped.")
        return 0

    # ---- BQ load + dedup per grain ----
    bq = bigquery.Client(project=project)
    _ensure_dataset(bq, "csm_staging")
    _ensure_dataset(bq, "csm_data")
    for grain, table in GRAIN_TABLE.items():
        _load_grain(bq, project, gcs_bucket, grain, table, full_cols)

    print("\nDONE.")
    return 0


def _ensure_dataset(bq: bigquery.Client, name: str) -> None:
    from google.cloud.exceptions import NotFound

    ds_id = f"{bq.project}.{name}"
    try:
        bq.get_dataset(ds_id)
    except NotFound:
        ds = bigquery.Dataset(ds_id)
        ds.location = "US"
        bq.create_dataset(ds, timeout=30)
        print(f"created dataset {ds_id}")


def _bq_schema(full_cols: list[str]) -> list:
    """Explicit BQ schema so cross-file Parquet type drift can't break the load
    (e.g. total_conversions written INT64 in one file, DOUBLE in another)."""
    fields = []
    for c in full_cols:
        if c in INT_COLS:
            t = "INT64"
        elif c in FLOAT_COLS:
            t = "FLOAT64"
        elif c == "date":
            t = "DATE"
        else:
            t = "STRING"
        fields.append(bigquery.SchemaField(c, t))
    return fields


def _load_grain(bq, project, gcs_bucket, grain, table, full_cols) -> None:
    src_uri = f"gs://{gcs_bucket}/{PARQUET_PREFIX}{grain}/*.parquet"
    stg = f"{project}.csm_staging.{table}__stg"
    prod = f"{project}.csm_data.{table}"

    # load all Parquet for this grain into staging (replace each run)
    job = bq.load_table_from_uri(
        src_uri,
        stg,
        job_config=bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.PARQUET,
            write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
            schema=_bq_schema(full_cols),
        ),
    )
    try:
        job.result()
    except Exception as e:  # noqa: BLE001
        print(f"[{grain}] no Parquet to load yet or load failed: {repr(e)[:160]}")
        return
    n_stg = bq.get_table(stg).num_rows
    print(f"[{grain}] staged {n_stg:,} rows -> {stg}")

    # de-dup into production: latest source file wins
    keys = DEDUP_KEYS[grain]
    part = ", ".join(keys)
    order = "source_filename DESC"  # filename embeds the YYYYMMDD_HHMMSS gen time
    # 'date' is already a DATE column, so partition on it directly (no DATE()).
    part_clause = (
        "PARTITION BY date CLUSTER BY source_account_id, campaign_id"
        if grain == "daily"
        else ""
    )
    sql = f"""
    CREATE OR REPLACE TABLE `{prod}`
    {part_clause}
    AS
    SELECT * EXCEPT(_rn) FROM (
      SELECT *, ROW_NUMBER() OVER (
        PARTITION BY {part} ORDER BY {order}
      ) AS _rn
      FROM `{stg}`
    )
    WHERE _rn = 1
    """
    bq.query(sql).result()
    n_prod = bq.get_table(prod).num_rows
    print(
        f"[{grain}] de-duped -> {prod}: {n_prod:,} rows "
        f"({n_stg - n_prod:,} duplicate rows collapsed)"
    )


if __name__ == "__main__":
    raise SystemExit(main())
