#!/usr/bin/env python3
"""
Read-only header scan for the email-delivered CM360 .xlsx corpus.

Scans every gs://email-file-ingestion/cm360/**/*.xlsx, locates each file's
header row (the preamble height varies file to file), and emits:

  - the definitive UNION of all distinct column names across the corpus
  - a per-variant manifest (distinct header signatures + example files)
  - the distribution of header-row offsets and "Date" column positions

This writes NOTHING to GCS or BigQuery and does not move any data. It only
downloads each .xlsx to a local scratch dir (first ~25 rows are enough, but
openpyxl needs the whole file) and reads the top of the sheet.

Usage:
    python scripts/cms_xlsx_header_scan.py [--limit N] [--out DIR]

Auth: uses GOOGLE_APPLICATION_CREDENTIALS from .env (the bionews SA that has
read access to the bucket).
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    print("python-dotenv not installed; run inside the project venv", file=sys.stderr)
    raise

import openpyxl
from google.cloud import storage

BUCKET = "email-file-ingestion"
PREFIX = "cm360/"
SHEET_CANDIDATES = ("Data",)  # all sampled files use a single sheet named 'Data'
SCAN_ROWS = 30  # header always lands within the first ~21 rows; 30 is safe headroom

# A row is the header row if it contains a "Date" cell AND a metric/dimension
# anchor. Keying on names (not position) tolerates the preamble-height drift and
# the fact that "Date" is not always in column 0.
ANCHOR_TOKENS = ("impression", "placement", "campaign")

# Path like: cm360/2021-04-12/795475_WeeklyDCMReport_..._20210412_043010_3267857841.xlsx
PATH_RE = re.compile(r"cm360/(\d{4}-\d{2}-\d{2})/(\d+)_(.+?)\.xlsx$", re.IGNORECASE)


def snake(name: str) -> str:
    """Snake_case a source header, e.g. 'Active View: Viewable Impressions'.

    Preserves a leading '%' as a 'pct_' token so that
    'Active View: % Viewable Impressions' (a rate) does NOT collide with
    'Active View: Viewable Impressions' (a count).
    """
    s = name.strip().lower()
    s = s.replace("(cm360)", "cm360").replace("(dcm)", "dcm")
    s = s.replace("%", " pct ")
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def find_header(rows: list[tuple]) -> tuple[int | None, list[str] | None]:
    for i, row in enumerate(rows):
        cells = ["" if c is None else str(c).strip() for c in row]
        nonempty = [c for c in cells if c]
        if not nonempty:
            continue
        has_date = "Date" in cells
        joined = " | ".join(nonempty).lower()
        is_anchor = any(tok in joined for tok in ANCHOR_TOKENS)
        # exclude the "MRC Accredited Metrics" / "Report Fields" preamble rows
        first = cells[0].lower()
        if has_date and is_anchor and "metric" not in first and "field" not in first:
            return i, nonempty
    return None, None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--limit", type=int, default=0, help="scan only first N files (0=all)"
    )
    ap.add_argument(
        "--out",
        default=os.path.join(
            os.environ.get("TEMP", "/tmp"),
            "claude",
            "c--orchestrator",
            "cms_header_scan",
        ),
        help="output dir for the manifest + downloaded files",
    )
    args = ap.parse_args()

    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    creds = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    project = os.getenv("GCP_PROJECT_ID", "bi-data-391216")
    print(f"creds: {creds}")
    print(f"project: {project}")

    os.makedirs(args.out, exist_ok=True)
    dl_dir = os.path.join(args.out, "dl")
    os.makedirs(dl_dir, exist_ok=True)

    client = storage.Client(project=project)
    bucket = client.bucket(BUCKET)

    blobs = [
        b
        for b in client.list_blobs(BUCKET, prefix=PREFIX)
        if b.name.lower().endswith(".xlsx")
    ]
    blobs.sort(key=lambda b: b.name)
    if args.limit:
        blobs = blobs[: args.limit]
    print(f"found {len(blobs)} .xlsx files under gs://{BUCKET}/{PREFIX}")

    union_raw: Counter = Counter()  # raw header name -> count of files
    union_snake: Counter = Counter()  # snake_cased name -> count of files
    sigs: dict[tuple, list[str]] = defaultdict(list)  # header tuple -> files
    header_row_dist: Counter = Counter()
    date_col_pos: Counter = Counter()
    sheet_names: Counter = Counter()
    accounts: Counter = Counter()
    errors: list[tuple[str, str]] = []
    no_header: list[str] = []

    for n, blob in enumerate(blobs, 1):
        local = os.path.join(dl_dir, os.path.basename(blob.name))
        try:
            if not os.path.exists(local):
                blob.download_to_filename(local)
            wb = openpyxl.load_workbook(local, read_only=True, data_only=True)
            sheet_names[tuple(wb.sheetnames)] += 1
            ws = wb[wb.sheetnames[0]]
            rows = []
            for i, r in enumerate(
                ws.iter_rows(min_row=1, max_row=SCAN_ROWS, values_only=True)
            ):
                rows.append(r)
            wb.close()
        except Exception as e:  # noqa: BLE001
            errors.append((blob.name, repr(e)[:160]))
            continue

        hr, cols = find_header(rows)
        if cols is None:
            no_header.append(blob.name)
            continue

        header_row_dist[hr] += 1
        date_col_pos[cols.index("Date") if "Date" in cols else -1] += 1
        sigs[tuple(cols)].append(blob.name)
        for c in cols:
            union_raw[c] += 1
            union_snake[snake(c)] += 1

        m = PATH_RE.search(blob.name)
        if m:
            accounts[m.group(2)] += 1

        if n % 200 == 0:
            print(f"  ...{n}/{len(blobs)} scanned")

    # ---- report ----
    print("\n" + "=" * 70)
    print(
        f"SCANNED: {len(blobs)}  |  parsed: {sum(len(v) for v in sigs.values())}"
        f"  |  no-header: {len(no_header)}  |  errors: {len(errors)}"
    )
    print(f"distinct header signatures: {len(sigs)}")
    print(f"sheet name sets: {dict(sheet_names)}")
    print(f"header-row offsets: {dict(sorted(header_row_dist.items()))}")
    print(f"'Date' column positions: {dict(sorted(date_col_pos.items()))}")
    print(f"distinct account prefixes: {len(accounts)}")

    print(
        f"\nUNION of raw column names ({len(union_raw)} distinct), by file frequency:"
    )
    for name, cnt in union_raw.most_common():
        print(f"  {cnt:5d}  {name!r:55s} -> {snake(name)}")

    if no_header:
        print(f"\nNO-HEADER files ({len(no_header)}):")
        for nm in no_header[:30]:
            print("  ", nm)
    if errors:
        print(f"\nERRORS ({len(errors)}):")
        for nm, e in errors[:30]:
            print("  ", nm, "::", e)

    # ---- persist machine-readable manifest ----
    manifest = {
        "files_scanned": len(blobs),
        "distinct_signatures": len(sigs),
        "union_raw": union_raw.most_common(),
        "union_snake": union_snake.most_common(),
        "header_row_offsets": dict(sorted(header_row_dist.items())),
        "date_col_positions": dict(sorted(date_col_pos.items())),
        "sheet_name_sets": {",".join(k): v for k, v in sheet_names.items()},
        "no_header_files": no_header,
        "errors": errors,
        "signatures": [
            {"columns": list(cols), "n_files": len(files), "example": files[0]}
            for cols, files in sorted(sigs.items(), key=lambda x: -len(x[1]))
        ],
    }
    out_path = os.path.join(args.out, "header_manifest.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)
    print(f"\nmanifest written: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
