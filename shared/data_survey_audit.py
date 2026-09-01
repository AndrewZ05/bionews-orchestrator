# -*- coding: utf-8 -*-
"""
LimeSurvey Data Audit Tool - Comprehensive Edition

Enhanced audit for lime_surveys_columnar_completed with:
- Column-level statistics
- Question-level analysis with subquestion text
- Likert scale numeric analysis for agreement questions
- Complete picture of all survey data on one screen

Usage:
    python shared/data_survey_audit.py limesurvey_data lime_surveys_columnar_completed \
        --credentials path/to/credentials.json --project bi-data-391216
"""

import argparse
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.cloud import bigquery
from google.oauth2 import service_account
import logging

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

def run_bigquery(client, query):
    return client.query(query).result()


def fetch_audit_data(result):
    audit_data = []
    for row in result:
        try:
            audit_value = row['audit_column']
            count_value = row['count']
        except (KeyError, TypeError):
            audit_value = row[0]
            count_value = row[1]
        audit_data.append((audit_value, count_value))
    return audit_data


def style_header_row(sheet, row_num, num_cols):
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font


def style_section_header(sheet, row_num, text, num_cols=6):
    """Apply section header styling."""
    sheet.append([text])
    cell = sheet.cell(row=row_num, column=1)
    cell.font = Font(bold=True, size=14)
    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


# ------------------------------------------------------------
# Likert Scale Analysis
# ------------------------------------------------------------

LIKERT_MAPPING = {
    'Strongly Agree': 5,
    'Agree': 4,
    'Somewhat Agree': 4,
    'Neutral': 3,
    'Somewhat Disagree': 2,
    'Disagree': 2,
    'Strongly Disagree': 1,
    'Does Not Apply': None,
}


def is_likert_question(answers):
    """Check if the answers suggest a Likert scale question."""
    likert_keywords = {'agree', 'disagree', 'neutral', 'strongly'}
    answer_words = set()
    for a in answers:
        if a:
            answer_words.update(a.lower().split())
    return len(likert_keywords & answer_words) >= 2


# ------------------------------------------------------------
# Comprehensive Question Analysis
# ------------------------------------------------------------

def audit_questions_comprehensive(bigquery_client, database_name, table_name, workbook):
    """Create a comprehensive question-by-question audit sheet."""

    print("  Creating comprehensive question analysis...", flush=True)

    sheet_name = "Question Analysis"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)

    full_table_name = f"`{database_name}.{table_name}`"

    # Get all questions with their stats
    questions_query = f'''
    SELECT
        standardized_question,
        COUNT(*) as total_rows,
        COUNT(DISTINCT response_id) as unique_respondents,
        COUNT(DISTINCT standardized_answer) as unique_answers,
        COUNTIF(subquestion_id IS NOT NULL AND TRIM(subquestion_id) != '') as rows_with_subq,
        COUNT(DISTINCT subquestion_id) as unique_subquestions
    FROM {full_table_name}
    WHERE standardized_question IS NOT NULL
    GROUP BY 1
    ORDER BY total_rows DESC
    '''

    # Get subquestion text mapping
    subq_query = f'''
    SELECT
        parent.qid as parent_qid,
        REGEXP_EXTRACT(sub.title, r"SQ(\\d+)") as subq_num,
        sub_l10n.question as subquestion_text
    FROM `{database_name}.lime_questions` parent
    JOIN `{database_name}.lime_questions` sub
        ON parent.qid = sub.parent_qid
    JOIN `{database_name}.lime_question_l10ns` sub_l10n
        ON sub.qid = sub_l10n.qid AND sub_l10n.language = "en"
    WHERE parent.parent_qid = 0
    '''

    subq_mapping = {}
    try:
        for row in run_bigquery(bigquery_client, subq_query):
            key = (row['parent_qid'], row['subq_num'])
            subq_mapping[key] = row['subquestion_text']
    except:
        pass

    # Header
    sheet.append(["COMPREHENSIVE QUESTION ANALYSIS"])
    sheet.append([f"Table: {table_name}"])
    sheet.append([""])

    row_num = 4

    for q_row in run_bigquery(bigquery_client, questions_query):
        question = q_row['standardized_question']
        has_subquestions = q_row['unique_subquestions'] > 1

        # Question header
        sheet.append(["=" * 100])
        sheet.append([f"QUESTION: {question}"])
        sheet.append([
            f"Total Rows: {q_row['total_rows']:,}",
            f"Respondents: {q_row['unique_respondents']:,}",
            f"Unique Answers: {q_row['unique_answers']:,}",
            f"Has Subquestions: {'Yes' if has_subquestions else 'No'}"
        ])
        sheet.append([""])

        if has_subquestions:
            # Get subquestion breakdown
            subq_detail_query = f'''
            WITH subq_mapping AS (
                SELECT
                    parent.qid as parent_qid,
                    REGEXP_EXTRACT(sub.title, r"SQ(\\d+)") as subq_num,
                    sub_l10n.question as subquestion_text
                FROM `{database_name}.lime_questions` parent
                JOIN `{database_name}.lime_questions` sub
                    ON parent.qid = sub.parent_qid
                JOIN `{database_name}.lime_question_l10ns` sub_l10n
                    ON sub.qid = sub_l10n.qid AND sub_l10n.language = "en"
                WHERE parent.parent_qid = 0
            )
            SELECT
                c.subquestion_id,
                COALESCE(sq.subquestion_text,
                    CASE WHEN c.subquestion_id = 'other' THEN '(Free-text other responses)'
                    ELSE 'Subquestion ' || c.subquestion_id END
                ) as subquestion_text,
                COUNT(*) as row_count,
                COUNT(DISTINCT c.standardized_answer) as unique_answers
            FROM {full_table_name} c
            LEFT JOIN subq_mapping sq
                ON c.question_id = sq.parent_qid
                AND c.subquestion_id = sq.subq_num
            WHERE c.standardized_question = "{question}"
              AND c.subquestion_id IS NOT NULL
              AND TRIM(c.subquestion_id) != ''
            GROUP BY 1, 2
            ORDER BY c.subquestion_id
            '''

            sheet.append(["Subquestions:"])
            header_row = sheet.max_row + 1
            sheet.append(["ID", "Subquestion Text", "Rows", "Unique Answers"])
            style_header_row(sheet, header_row, 4)

            for subq in run_bigquery(bigquery_client, subq_detail_query):
                text = subq['subquestion_text']
                if text and len(text) > 80:
                    text = text[:77] + "..."
                sheet.append([
                    subq['subquestion_id'],
                    text,
                    int(subq['row_count']),
                    int(subq['unique_answers'])
                ])

            sheet.append([""])

        # Get answer distribution
        answer_query = f'''
        SELECT
            standardized_answer,
            COUNT(*) as cnt
        FROM {full_table_name}
        WHERE standardized_question = "{question}"
          AND standardized_answer IS NOT NULL
        GROUP BY 1
        ORDER BY cnt DESC
        LIMIT 25
        '''

        answers = list(run_bigquery(bigquery_client, answer_query))
        answer_texts = [a['standardized_answer'] for a in answers]

        # Check if this is a Likert scale question
        if is_likert_question(answer_texts):
            sheet.append(["Answer Distribution (Likert Scale):"])
            header_row = sheet.max_row + 1
            sheet.append(["Answer", "Numeric Value", "Count", "Percentage"])
            style_header_row(sheet, header_row, 4)

            total = sum(a['cnt'] for a in answers)
            for a in answers:
                answer = a['standardized_answer']
                numeric = LIKERT_MAPPING.get(answer, '-')
                pct = (a['cnt'] / total * 100) if total > 0 else 0
                sheet.append([
                    answer,
                    numeric if numeric else 'N/A',
                    int(a['cnt']),
                    f"{pct:.1f}%"
                ])

            # Calculate summary stats for Likert
            numeric_answers = [(a['standardized_answer'], a['cnt']) for a in answers
                              if LIKERT_MAPPING.get(a['standardized_answer'])]
            if numeric_answers:
                total_numeric = sum(cnt for _, cnt in numeric_answers)
                weighted_sum = sum(LIKERT_MAPPING[ans] * cnt for ans, cnt in numeric_answers)
                avg_score = weighted_sum / total_numeric if total_numeric > 0 else 0

                agree_count = sum(cnt for ans, cnt in numeric_answers if LIKERT_MAPPING.get(ans, 0) >= 4)
                disagree_count = sum(cnt for ans, cnt in numeric_answers if LIKERT_MAPPING.get(ans, 0) <= 2)

                sheet.append([""])
                sheet.append([f"Likert Summary: Avg={avg_score:.2f}, Agree={agree_count:,} ({agree_count/total_numeric*100:.1f}%), Disagree={disagree_count:,} ({disagree_count/total_numeric*100:.1f}%)"])
        else:
            # Regular answer distribution
            sheet.append(["Answer Distribution:"])
            header_row = sheet.max_row + 1
            sheet.append(["Answer", "Count", "Percentage"])
            style_header_row(sheet, header_row, 3)

            total = sum(a['cnt'] for a in answers)
            for a in answers:
                answer = a['standardized_answer']
                if answer and len(answer) > 60:
                    answer = answer[:57] + "..."
                pct = (a['cnt'] / total * 100) if total > 0 else 0
                sheet.append([
                    answer,
                    int(a['cnt']),
                    f"{pct:.1f}%"
                ])

        sheet.append([""])
        sheet.append([""])

    # Adjust column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 80
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 15

    print(f"    Created Question Analysis sheet", flush=True)


# ------------------------------------------------------------
# Data Quality Summary
# ------------------------------------------------------------

def audit_data_quality_summary(bigquery_client, database_name, table_name, workbook):
    """Create a data quality summary sheet."""

    print("  Creating data quality summary...", flush=True)

    sheet_name = "Data Quality"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name, 0)  # Insert at beginning

    full_table_name = f"`{database_name}.{table_name}`"

    # Overall stats
    stats_query = f'''
    SELECT
        COUNT(*) as total_rows,
        COUNT(DISTINCT response_id) as unique_responses,
        COUNT(DISTINCT survey_id) as unique_surveys,
        COUNT(DISTINCT standardized_question) as unique_questions,
        COUNT(DISTINCT site) as unique_sites,
        MIN(submitdate) as earliest_date,
        MAX(submitdate) as latest_date,
        COUNTIF(standardized_answer IS NULL) as null_answers,
        COUNTIF(subquestion_id IS NOT NULL) as rows_with_subq
    FROM {full_table_name}
    '''

    stats = list(run_bigquery(bigquery_client, stats_query))[0]

    sheet.append(["LIMESURVEY DATA QUALITY SUMMARY"])
    sheet.append([f"Table: {table_name}"])
    sheet.append([""])

    sheet.append(["OVERVIEW"])
    sheet.append(["Metric", "Value"])
    sheet.append(["Total Rows", f"{stats['total_rows']:,}"])
    sheet.append(["Unique Responses", f"{stats['unique_responses']:,}"])
    sheet.append(["Unique Surveys", f"{stats['unique_surveys']:,}"])
    sheet.append(["Unique Questions", f"{stats['unique_questions']:,}"])
    sheet.append(["Unique Sites", f"{stats['unique_sites']:,}"])
    sheet.append(["Date Range", f"{stats['earliest_date']} to {stats['latest_date']}"])
    sheet.append(["Rows with Subquestions", f"{stats['rows_with_subq']:,} ({stats['rows_with_subq']/stats['total_rows']*100:.1f}%)"])
    sheet.append(["Null Answers", f"{stats['null_answers']:,} ({stats['null_answers']/stats['total_rows']*100:.1f}%)"])
    sheet.append([""])

    # Questions by type
    sheet.append(["QUESTIONS BY TYPE"])

    type_query = f'''
    SELECT
        standardized_question,
        COUNT(*) as rows,
        COUNT(DISTINCT response_id) as respondents,
        COUNT(DISTINCT standardized_answer) as unique_answers,
        COUNT(DISTINCT subquestion_id) as subquestion_count,
        CASE
            WHEN COUNT(DISTINCT subquestion_id) > 1 THEN 'Matrix/Grid'
            WHEN COUNT(DISTINCT standardized_answer) <= 10 THEN 'Single Choice'
            WHEN COUNT(DISTINCT standardized_answer) <= 50 THEN 'Multiple Choice'
            ELSE 'Free Text'
        END as question_type
    FROM {full_table_name}
    WHERE standardized_question IS NOT NULL
    GROUP BY 1
    ORDER BY rows DESC
    '''

    header_row = sheet.max_row + 1
    sheet.append(["Question", "Type", "Rows", "Respondents", "Answers", "Subquestions"])
    style_header_row(sheet, header_row, 6)

    for row in run_bigquery(bigquery_client, type_query):
        q = row['standardized_question']
        if len(q) > 40:
            q = q[:37] + "..."
        sheet.append([
            q,
            row['question_type'],
            int(row['rows']),
            int(row['respondents']),
            int(row['unique_answers']),
            int(row['subquestion_count']) if row['subquestion_count'] > 1 else '-'
        ])

    sheet.append([""])

    # Potential data issues
    sheet.append(["POTENTIAL DATA ISSUES"])

    issues_query = f'''
    SELECT
        'Long answers (>200 chars)' as issue,
        COUNT(*) as count
    FROM {full_table_name}
    WHERE LENGTH(standardized_answer) > 200

    UNION ALL

    SELECT
        'UMLS corruption (chemical names)' as issue,
        COUNT(*) as count
    FROM {full_table_name}
    WHERE standardized_answer LIKE '%dipalmitoyl%'
       OR standardized_answer LIKE '%tetrahydropyridine%'
       OR standardized_answer LIKE '%Glucan%'
       OR standardized_answer LIKE '%Acute Physiology%'

    UNION ALL

    SELECT
        'Undecoded answer codes (AO01, etc.)' as issue,
        COUNT(*) as count
    FROM {full_table_name}
    WHERE REGEXP_CONTAINS(standardized_answer, r'^AO[0-9]{{2}}$')

    UNION ALL

    SELECT
        'Empty standardized_question' as issue,
        COUNT(*) as count
    FROM {full_table_name}
    WHERE standardized_question IS NULL OR TRIM(standardized_question) = ''
    '''

    header_row = sheet.max_row + 1
    sheet.append(["Issue", "Count", "Status"])
    style_header_row(sheet, header_row, 3)

    for row in run_bigquery(bigquery_client, issues_query):
        status = "OK" if row['count'] == 0 else "REVIEW"
        sheet.append([row['issue'], int(row['count']), status])

    # Adjust column widths
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12

    print(f"    Created Data Quality sheet", flush=True)


# ------------------------------------------------------------
# Core audit (standard column audit)
# ------------------------------------------------------------

def audit_table(bigquery_client, database_name, table_name, workbook, worksheet_name):
    print(f'Auditing table: {table_name} in {database_name}', flush=True)

    full_table_name = f"`{database_name}.{table_name}`"

    if worksheet_name not in workbook.sheetnames:
        workbook.create_sheet(worksheet_name)
    sheet = workbook[worksheet_name]

    # --------------------------------------------------------
    # Row count
    # --------------------------------------------------------
    total_rows = list(
        run_bigquery(bigquery_client, f"SELECT COUNT(*) FROM {full_table_name}")
    )[0][0]

    # --------------------------------------------------------
    # Column metadata
    # --------------------------------------------------------
    column_meta = {
        row["column_name"]: row["data_type"]
        for row in run_bigquery(
            bigquery_client,
            f"""
            SELECT column_name, data_type
            FROM `{database_name}.INFORMATION_SCHEMA.COLUMNS`
            WHERE table_name = '{table_name}'
            ORDER BY ordinal_position
            """
        )
    }

    if not column_meta:
        print(f"Warning: No columns found for table {table_name}")
        return

    sheet.append([f"Table name: {table_name}"])
    sheet.append([f"Number of rows: {total_rows:,}"])
    sheet.append([""])

    # --------------------------------------------------------
    # Per-column audit (abbreviated for key columns only)
    # --------------------------------------------------------
    key_columns = ['standardized_question', 'standardized_answer', 'subquestion_id',
                   'response_id', 'survey_id', 'site', 'submitdate']

    for column_name, data_type in column_meta.items():
        # Only detailed audit for key columns, brief for others
        if column_name not in key_columns and 'lime_surveys_columnar' in table_name:
            continue

        print(f"  Auditing column: {column_name}", flush=True)

        try:
            escaped_column = f"`{column_name}`"

            # Max length
            max_length = list(run_bigquery(
                bigquery_client,
                f"""
                SELECT MAX(LENGTH(CAST({escaped_column} AS STRING)))
                FROM {full_table_name}
                WHERE {escaped_column} IS NOT NULL
                """
            ))[0][0] or 0

            # Non-null count
            non_null_count = list(run_bigquery(
                bigquery_client,
                f"""
                SELECT COUNT(*)
                FROM {full_table_name}
                WHERE CAST({escaped_column} AS STRING) > ''
                """
            ))[0][0]

            # Distinct count
            distinct_count = list(run_bigquery(
                bigquery_client,
                f"""
                SELECT COUNT(DISTINCT {escaped_column})
                FROM {full_table_name}
                """
            ))[0][0]

            fill_rate = (non_null_count / total_rows * 100) if total_rows else 0
            null_count = total_rows - non_null_count
            uniqueness_ratio = (distinct_count / non_null_count * 100) if non_null_count else 0

            sheet.append([f"Column Name: {column_name} ({data_type})"])
            sheet.append([f"Maximum length: {max_length} | Fill rate: {fill_rate:.2f}% | Distinct: {distinct_count:,}"])

            # Value distributions (top 15 only)
            if distinct_count <= 100:
                limit = min(distinct_count, 30)
            else:
                limit = 15

            values = fetch_audit_data(run_bigquery(
                bigquery_client,
                f"""
                SELECT CAST({escaped_column} AS STRING) AS audit_column, COUNT(*) AS count
                FROM {full_table_name}
                WHERE {escaped_column} IS NOT NULL
                GROUP BY {escaped_column}
                ORDER BY count DESC
                LIMIT {limit}
                """
            ))

            sheet.append(["Value", "Count"])
            for value, count in values:
                display_value = str(value)[:80] if value else "(empty)"
                sheet.append([display_value, int(count)])

            sheet.append([""])
            sheet.append(["----------------------------------------"])
            sheet.append([""])

        except Exception as e:
            print(f"    Error auditing column '{column_name}': {e}")
            sheet.append([f"Error auditing column {column_name}: {e}"])
            sheet.append([""])

    # --------------------------------------------------------
    # Survey-specific comprehensive analysis
    # --------------------------------------------------------
    if 'lime_surveys_columnar' in table_name:
        audit_data_quality_summary(bigquery_client, database_name, table_name, workbook)
        audit_questions_comprehensive(bigquery_client, database_name, table_name, workbook)


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------

def main(database_name, table_names, spreadsheet_name, credentials_path, project_id=None):
    try:
        credentials = service_account.Credentials.from_service_account_file(credentials_path)
        client = bigquery.Client(credentials=credentials, project=project_id)

        # Always create fresh workbook (overwrites existing file)
        workbook = Workbook()

        if 'Sheet' in workbook.sheetnames:
            workbook['Sheet'].title = 'Main'

        for table_name in table_names:
            audit_table(
                client,
                database_name,
                table_name,
                workbook,
                table_name[:31]
            )

        workbook.save(spreadsheet_name)
        print(f"\nAudit results written to {spreadsheet_name}", flush=True)

    except Exception as e:
        print("Data Audit could not be completed.")
        print(f"Error: {e}")
        raise
    finally:
        logging.shutdown()


# ------------------------------------------------------------
# CLI
# ------------------------------------------------------------

def quick_audit_question(client, database_name, table_name, question_name):
    """Quick console audit for a single standardized_question value."""
    full_table = f"`{database_name}.{table_name}`"

    # Get question stats
    stats_query = f'''
    SELECT
        COUNT(*) as total_rows,
        COUNT(DISTINCT response_id) as unique_respondents,
        COUNT(DISTINCT standardized_answer) as unique_answers,
        COUNT(DISTINCT subquestion_id) as unique_subquestions,
        MAX(LENGTH(standardized_answer)) as max_answer_len
    FROM {full_table}
    WHERE standardized_question = "{question_name}"
    '''

    stats = list(client.query(stats_query).result())[0]

    print(f"\nColumn Name: {question_name} (STRING)")
    print(f"Maximum length: {stats.max_answer_len} | Fill rate: 100% | Distinct: {stats.unique_answers}")
    print(f"Respondents: {stats.unique_respondents:,} | Subquestions: {stats.unique_subquestions}")
    print("Value".ljust(70) + "Count".rjust(10))
    print("-" * 80)

    # Get answer distribution
    answer_query = f'''
    SELECT standardized_answer, COUNT(*) as cnt
    FROM {full_table}
    WHERE standardized_question = "{question_name}"
      AND standardized_answer IS NOT NULL
    GROUP BY 1
    ORDER BY cnt DESC
    LIMIT 50
    '''

    for row in client.query(answer_query).result():
        ans = str(row.standardized_answer) if row.standardized_answer else "(null)"
        if len(ans) > 68:
            ans = ans[:65] + "..."
        print(f"{ans.ljust(70)}{row.cnt:>10,}")

    # If has subquestions, show breakdown
    if stats.unique_subquestions > 1:
        print(f"\n--- Subquestion Breakdown ---")
        subq_query = f'''
        SELECT
            subquestion_id,
            COUNT(*) as cnt,
            COUNT(DISTINCT standardized_answer) as unique_answers
        FROM {full_table}
        WHERE standardized_question = "{question_name}"
          AND subquestion_id IS NOT NULL AND subquestion_id != ''
        GROUP BY 1
        ORDER BY subquestion_id
        '''
        print("SubQ ID".ljust(10) + "Rows".rjust(10) + "Unique Answers".rjust(15))
        for row in client.query(subq_query).result():
            print(f"{str(row.subquestion_id).ljust(10)}{row.cnt:>10,}{row.unique_answers:>15,}")


def quick_audit_column(client, database_name, table_name, column_name):
    """Quick console audit for a column in a wide table."""
    full_table = f"`{database_name}.{table_name}`"

    # Get column stats
    stats_query = f'''
    SELECT
        COUNT(*) as total_rows,
        COUNTIF(`{column_name}` IS NOT NULL AND CAST(`{column_name}` AS STRING) != '') as non_null_rows,
        COUNT(DISTINCT `{column_name}`) as unique_values,
        MAX(LENGTH(CAST(`{column_name}` AS STRING))) as max_len
    FROM {full_table}
    '''

    stats = list(client.query(stats_query).result())[0]
    fill_rate = (stats.non_null_rows / stats.total_rows * 100) if stats.total_rows else 0

    # Get data type
    dtype_query = f'''
    SELECT data_type
    FROM `{database_name}.INFORMATION_SCHEMA.COLUMNS`
    WHERE table_name = "{table_name}" AND column_name = "{column_name}"
    '''
    dtype = list(client.query(dtype_query).result())[0].data_type

    print(f"\nColumn Name: {column_name} ({dtype})")
    print(f"Maximum length: {stats.max_len} | Fill rate: {fill_rate:.2f}% | Distinct: {stats.unique_values}")
    print("Value".ljust(70) + "Count".rjust(10))
    print("-" * 80)

    # Get value distribution
    value_query = f'''
    SELECT CAST(`{column_name}` AS STRING) as val, COUNT(*) as cnt
    FROM {full_table}
    WHERE `{column_name}` IS NOT NULL
    GROUP BY 1
    ORDER BY cnt DESC
    LIMIT 50
    '''

    for row in client.query(value_query).result():
        val = str(row.val) if row.val else "(null)"
        if len(val) > 68:
            val = val[:65] + "..."
        print(f"{val.ljust(70)}{row.cnt:>10,}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='LimeSurvey Data Audit Tool - Comprehensive Edition',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # Full audit to Excel
  python shared/data_survey_audit.py limesurvey_data lime_surveys_columnar_completed \\
      --credentials credentials.json --project bi-data-391216

  # Quick audit of a single question (console output)
  python shared/data_survey_audit.py limesurvey_data lime_surveys_columnar_completed \\
      --credentials credentials.json --project bi-data-391216 --question age

  # Quick audit with partial match
  python shared/data_survey_audit.py limesurvey_data lime_surveys_columnar_completed \\
      --credentials credentials.json --project bi-data-391216 --question treatment

Output Sheets (full audit):
  - Data Quality: Overall summary, question types, data issues
  - Question Analysis: Each question with subquestions and answer distributions
  - [table_name]: Column-level statistics
        '''
    )
    parser.add_argument('database_name', help='BigQuery dataset name (e.g., limesurvey_data)')
    parser.add_argument('table_names', nargs='+', help='Table name(s) to audit')
    parser.add_argument('--spreadsheet', default='survey_audit_results.xlsx', help='Output Excel file (always overwrites)')
    parser.add_argument('--credentials', required=True, help='Path to service account JSON')
    parser.add_argument('--project', help='GCP project ID')
    parser.add_argument('--question', '-q', help='Quick audit single question (prints to console)')

    args = parser.parse_args()

    if args.question:
        # Quick single-question/column audit to console
        credentials = service_account.Credentials.from_service_account_file(args.credentials)
        client = bigquery.Client(credentials=credentials, project=args.project)

        table_name = args.table_names[0]
        question = args.question

        # First check if table has standardized_question column (columnar format)
        schema_query = f'''
        SELECT column_name
        FROM `{args.database_name}.INFORMATION_SCHEMA.COLUMNS`
        WHERE table_name = "{table_name}" AND column_name = "standardized_question"
        '''
        has_std_question = len(list(client.query(schema_query).result())) > 0

        if has_std_question:
            # Columnar table - search by standardized_question value
            check_query = f'''
            SELECT DISTINCT standardized_question
            FROM `{args.database_name}.{table_name}`
            WHERE standardized_question = "{question}"
               OR standardized_question LIKE "%{question}%"
            ORDER BY standardized_question
            '''
            matches = [row.standardized_question for row in client.query(check_query).result()]

            if not matches:
                print(f"No questions found matching '{question}'")
            elif len(matches) == 1:
                quick_audit_question(client, args.database_name, table_name, matches[0])
            else:
                # If exact match exists, use it; otherwise show all matches
                if question in matches:
                    quick_audit_question(client, args.database_name, table_name, question)
                else:
                    print(f"Multiple questions match '{question}':")
                    for m in matches:
                        print(f"  - {m}")
                    print(f"\nShowing first match: {matches[0]}\n")
                    quick_audit_question(client, args.database_name, table_name, matches[0])
        else:
            # Wide table - search by column name
            col_query = f'''
            SELECT column_name, data_type
            FROM `{args.database_name}.INFORMATION_SCHEMA.COLUMNS`
            WHERE table_name = "{table_name}"
              AND (column_name = "{question}" OR column_name LIKE "%{question}%")
            ORDER BY column_name
            '''
            matches = [(row.column_name, row.data_type) for row in client.query(col_query).result()]

            if not matches:
                print(f"No columns found matching '{question}'")
            else:
                if len(matches) > 1 and question not in [m[0] for m in matches]:
                    print(f"Multiple columns match '{question}':")
                    for col, dtype in matches:
                        print(f"  - {col} ({dtype})")
                    print(f"\nShowing first match: {matches[0][0]}\n")

                col_name = question if question in [m[0] for m in matches] else matches[0][0]
                quick_audit_column(client, args.database_name, table_name, col_name)
    else:
        # Full audit to Excel (always overwrites existing file)
        main(
            database_name=args.database_name,
            table_names=args.table_names,
            spreadsheet_name=args.spreadsheet,
            credentials_path=args.credentials,
            project_id=args.project
        )
