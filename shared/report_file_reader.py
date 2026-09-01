#!/usr/bin/env python3
"""Shared reader for email-delivered ad-ops report files living in GCS.

WHY
  DoubleVerify and Trade Desk arrive as vendor report files dropped into
  gs://email-file-ingestion/<source>/<YYYY-MM-DD>/*. Every other ETL source in
  this repo talks to an API; these two read files. This module is the common
  file-handling layer so plugins/doubleverify_extractor.py and
  plugins/tradedesk_extractor.py stay thin BaseExtractor subclasses.

  The logic here is ported from scripts/adops_email_backfill.py (a standalone
  script that loaded these sources ONCE, on 2026-06-25, and was never
  scheduled). What is NOT ported: that script's per-file Parquet staging,
  wide-union column discovery, and its own BigQuery load/dedup half. Those are
  exactly what the plugin architecture already provides -- the plugin returns
  records, orchestrate.py does GCS upload -> external table -> staging -> hash
  merge.

KEY BEHAVIORS
  - Routing is CONFIG-DRIVEN: each resource in the source YAML declares a
    file_match.filename_regex, so a new vendor report variant is a config
    change, not a code change. A file matching nothing is reported as
    "unrouted" -- the only signal that a new variant appeared.
  - Typing is driven by the resource's YAML `schema:` block, so the YAML is the
    single source of truth. (The origin script kept parallel int_cols/float_cols
    sets, a second place to forget to update.)
  - dedup_latest_wins() is MANDATORY before returning records. These feeds
    restate: DV files carry a full month-to-date window, TTD files a full year.
    A single lookback window therefore pulls the same (date, dims) row from many
    files. BigQuery MERGE rejects a source with duplicate key rows outright
    ("UPDATE/MERGE must match at most one source row").
  - coalesce_key_columns() fills nullable KEY columns with ''. The hash merge
    joins with `main.k = staging.k` (shared/bigquery_utils.py:560) and has no
    NULL guard. In BigQuery NULL = NULL is NULL, not TRUE -- so a NULL key never
    matches an existing row and would re-INSERT it on every run, growing the
    table without bound.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# Folder layout inside the ingestion bucket: <source_prefix>/<YYYY-MM-DD>/<file>
DAY_FOLDER_RE = re.compile(r"/(\d{4}-\d{2}-\d{2})/")

# Lineage columns stamped onto every row. Declared in each resource's YAML
# schema and listed in hash_merge.exclude_from_hash -- they identify WHERE a row
# came from and must not perturb the row hash, or every restated row would look
# changed on every run.
LINEAGE_COLUMNS = ("source_filename", "source_gcs_path", "report_date_partition")


# --------------------------------------------------------------------------- #
# Column-name and content helpers (ported from scripts/adops_email_backfill.py)
# --------------------------------------------------------------------------- #
def snake(name: str) -> str:
    """Normalize a vendor header to a BigQuery-legal snake_case column name.

    Ported from scripts/adops_email_backfill.py:186, with one addition: a
    BigQuery column may not START with a digit. The DoubleVerify viewability
    report has 8 such headers ("100% Display Viewable Impressions",
    "50% Display Viewable 1-5 Secs Rate", ...) which the original rules turned
    into "100_pct_display_viewable_impressions" -- an illegal column name that
    fails at table creation. Those get a "pct_" prefix. Every other header is
    unchanged, so names stay byte-identical to the original backfill.
    """
    s = name.strip().lower()
    s = s.replace("%", " pct ").replace("/", " ").replace("(", " ").replace(")", " ")
    s = re.sub(r"[^\w]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if s and s[0].isdigit():
        s = f"pct_{s}"
    return s


def is_html(raw: bytes) -> bool:
    """True when the payload is an HTML page rather than a report.

    The DoubleVerify email feed captured its own login page instead of data for
    every day from 2025-02-01 to 2025-06-13 (~318 files). Those are detected and
    skipped, never loaded. Ported from scripts/adops_email_backfill.py:194.
    """
    head = raw[:300].decode("utf-8-sig", "replace").lstrip().lower()
    return head.startswith("<!doctype") or "<html" in head[:200]


def read_source_file(payload: bytes, reader: str) -> Optional[pd.DataFrame]:
    """Parse raw file bytes into a DataFrame of objects (no type coercion yet).

    reader:
      "csv"             -- DoubleVerify
      "xlsx_data_sheet" -- Trade Desk. The workbook carries TWO sheets: the real
                           data in "Ad Group Performance_data" and an empty
                           11x7 formatting shell named "Ad Group Performance".
                           Selecting the wrong one silently yields no rows, so
                           the "_data" suffix match is load-bearing and an
                           empty result raises rather than returning the shell.
    """
    if reader == "csv":
        return pd.read_csv(
            io.BytesIO(payload),
            dtype=object,
            encoding="utf-8-sig",
            on_bad_lines="skip",
        )

    if reader == "xlsx_data_sheet":
        buf = io.BytesIO(payload)
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        data_sheets = [s for s in wb.sheetnames if s.endswith("_data")]
        sheet = data_sheets[0] if data_sheets else wb.sheetnames[0]
        wb.close()

        df = pd.read_excel(
            io.BytesIO(payload), sheet_name=sheet, dtype=object, engine="openpyxl"
        )
        if df is None or df.dropna(how="all").empty:
            # Reading the formatting shell instead of the data sheet is a silent
            # data-loss bug -- fail loudly so the run records it.
            raise ValueError(
                f"xlsx sheet '{sheet}' produced no rows (expected a '*_data' "
                f"sheet; available: {data_sheets or 'none'})"
            )
        return df

    raise ValueError(f"unknown reader {reader!r}")


# --------------------------------------------------------------------------- #
# GCS listing, scoped to the day folders in the extraction window
# --------------------------------------------------------------------------- #
def _parse_day(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def day_folder_of(blob_name: str) -> Optional[str]:
    """Return the '<YYYY-MM-DD>' folder a blob sits in, or None."""
    m = DAY_FOLDER_RE.search(
        f"/{blob_name}" if not blob_name.startswith("/") else blob_name
    )
    return m.group(1) if m else None


def list_day_folder_blobs(
    client: Any,
    bucket: str,
    prefix: str,
    start_date: Optional[str],
    end_date: Optional[str],
    ext: str,
) -> List[Any]:
    """List source blobs, restricted to the day folders in [start_date, end_date].

    A nightly run touches ~7 day folders, so this issues a handful of tightly
    scoped prefix listings instead of scanning all ~1,600 objects. When no start
    date is given (refresh_mode=full) it falls back to listing the whole prefix.
    """
    ext = ext.lower()
    start = _parse_day(start_date)
    end = _parse_day(end_date) or date.today()

    if start is None:
        logger.info("No start date -- listing full prefix gs://%s/%s", bucket, prefix)
        blobs = [
            b
            for b in client.list_blobs(bucket, prefix=prefix)
            if b.name.lower().endswith(ext)
        ]
        blobs.sort(key=lambda b: b.name)
        return blobs

    if end < start:
        start, end = end, start

    blobs: List[Any] = []
    day = start
    while day <= end:
        day_prefix = f"{prefix.rstrip('/')}/{day.isoformat()}/"
        blobs.extend(
            b
            for b in client.list_blobs(bucket, prefix=day_prefix)
            if b.name.lower().endswith(ext)
        )
        day += timedelta(days=1)

    blobs.sort(key=lambda b: b.name)
    logger.info(
        "Listed %d %s files across %d day folders (%s -> %s)",
        len(blobs),
        ext,
        (end - start).days + 1,
        start,
        end,
    )
    return blobs


# --------------------------------------------------------------------------- #
# Config-driven file -> resource routing
# --------------------------------------------------------------------------- #
class ResourceSpec:
    """One resource's file_match rules, compiled from the source YAML."""

    __slots__ = (
        "name",
        "regex",
        "reader",
        "extension",
        "priority",
        "required_columns",
        "html_expected_before",
    )

    def __init__(self, name: str, file_match: Dict[str, Any]):
        self.name = name
        pattern = file_match.get("filename_regex") or ""
        self.regex = re.compile(pattern, re.IGNORECASE) if pattern else None
        self.reader = file_match.get("reader", "csv")
        self.extension = str(file_match.get("extension", ".csv")).lower()
        # Higher priority wins when several regexes match the same filename.
        self.priority = int(file_match.get("priority", 0))
        self.required_columns = list(file_match.get("required_columns") or [])
        self.html_expected_before = file_match.get("html_expected_before")

    def matches(self, blob_name: str) -> bool:
        if self.regex is None:
            return False
        if not blob_name.lower().endswith(self.extension):
            return False
        return bool(self.regex.search(blob_name.rsplit("/", 1)[-1]))


def build_resource_specs(config: Dict[str, Any]) -> List[ResourceSpec]:
    """Compile every ACTIVE resource's file_match block, highest priority first."""
    specs: List[ResourceSpec] = []
    for name, rc in (config.get("resources") or {}).items():
        if rc.get("active") is False:
            continue
        fm = rc.get("file_match")
        if not fm:
            continue
        specs.append(ResourceSpec(name, fm))
    specs.sort(key=lambda s: -s.priority)
    return specs


def route_blob(blob_name: str, specs: Sequence[ResourceSpec]) -> Optional[str]:
    """Return the logical resource a file belongs to, or None if unrouted.

    First match in priority order wins. None means a report variant arrived that
    no resource claims -- callers surface that as a WARNING, because it is the
    only signal that the vendor added a new report.
    """
    for spec in specs:
        if spec.matches(blob_name):
            return spec.name
    return None


# --------------------------------------------------------------------------- #
# Typing + normalization, driven by the resource's YAML schema block
# --------------------------------------------------------------------------- #
_INT_TYPES = {"INT64", "INTEGER", "INT"}
_FLOAT_TYPES = {"FLOAT64", "FLOAT", "NUMERIC", "BIGNUMERIC"}
_DATE_TYPES = {"DATE"}
_TIMESTAMP_TYPES = {"TIMESTAMP", "DATETIME"}


def normalize_frame(df: pd.DataFrame, schema_map: Dict[str, str]) -> pd.DataFrame:
    """snake_case the headers, drop junk columns/rows, and coerce to YAML types.

    Columns are reindexed to the YAML schema, so a source file missing a column
    yields NULLs for it and an unexpected extra column is dropped rather than
    silently widening the table. Lineage/system columns are left for
    stamp_lineage() and the extraction accumulator to fill in.
    """
    df = df.copy()
    df.columns = [snake(str(c)) for c in df.columns]

    # openpyxl/pandas name blank columns "unnamed_0" etc.
    df = df.loc[:, [c for c in df.columns if c and not c.startswith("unnamed")]]
    # Collapse duplicate headers (keep first) so reindex cannot explode.
    df = df.loc[:, ~pd.Index(df.columns).duplicated()]

    # Vendor exports often carry blank trailing/total rows.
    if len(df.columns):
        first = df.iloc[:, 0].astype(str)
        df = df[
            first.notna() & (first.str.strip() != "") & (first.str.lower() != "nan")
        ]

    data_cols = [
        c
        for c in schema_map
        if c not in LINEAGE_COLUMNS
        and c not in ("extracted_at", "execution_id", "source")
    ]
    df = df.reindex(columns=data_cols)

    for col in df.columns:
        bq_type = str(schema_map.get(col, "STRING")).upper()
        if bq_type in _INT_TYPES:
            df[col] = pd.to_numeric(df[col], errors="coerce").round().astype("Int64")
        elif bq_type in _FLOAT_TYPES:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("float64")
        elif bq_type in _DATE_TYPES:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
        elif bq_type in _TIMESTAMP_TYPES:
            df[col] = pd.to_datetime(df[col], errors="coerce")
        else:
            df[col] = df[col].astype("string")

    return df


def stamp_lineage(df: pd.DataFrame, blob_name: str, bucket: str) -> pd.DataFrame:
    """Add source_filename / source_gcs_path / report_date_partition.

    report_date_partition is a real DATE (the origin script left it a string).
    For the one DoubleVerify report that has no Date column at all, this IS the
    partition field and part of the primary key, so its type matters.
    """
    df = df.copy()
    folder = day_folder_of(blob_name)
    df["source_filename"] = blob_name.rsplit("/", 1)[-1]
    df["source_gcs_path"] = f"gs://{bucket}/{blob_name}"
    df["report_date_partition"] = pd.to_datetime(folder).date() if folder else None
    return df


def assert_required_columns(
    df: pd.DataFrame, spec: ResourceSpec, blob_name: str
) -> None:
    """Fail loudly when a routed file does not look like the expected report.

    Defense in depth against a filename regex matching the wrong file: without
    this, a mis-routed file writes a full set of NULLs into the target table.
    """
    if not spec.required_columns:
        return
    missing = [c for c in spec.required_columns if c not in df.columns]
    if missing:
        raise ValueError(
            f"{blob_name} routed to '{spec.name}' but is missing required "
            f"columns {missing} -- filename_regex likely matches the wrong report"
        )


def coalesce_key_columns(df: pd.DataFrame, key_cols: Iterable[str]) -> pd.DataFrame:
    """Replace NULLs in KEY columns with '' so the hash merge can join on them.

    shared/bigquery_utils.py:560 builds the MERGE predicate as
    `main.k = staging.k` with no NULL guard. Because NULL = NULL evaluates to
    NULL (not TRUE) in BigQuery, a NULL key never matches its existing row: the
    MERGE takes the NOT MATCHED branch and INSERTs a duplicate on every single
    run. Filling with '' keeps the join total.

    Concretely: dv_brand_detail merges the 40-column and 42-column DoubleVerify
    brand reports. placement_size/ad_size exist only in the 42-column variant,
    so they are NULL for 310 of 372 files and are part of the key.
    """
    df = df.copy()
    for col in key_cols:
        if col not in df.columns:
            continue
        if str(df[col].dtype) in ("object", "string"):
            df[col] = df[col].astype("string").fillna("")
        else:
            # Non-string key columns (dates, ints) cannot take '' -- a NULL there
            # means the key itself is wrong, so surface it instead of hiding it.
            null_count = int(df[col].isna().sum())
            if null_count:
                logger.warning(
                    "Key column '%s' has %d NULL values but is typed %s -- these "
                    "rows will re-insert on every run; the key may be wrong",
                    col,
                    null_count,
                    df[col].dtype,
                )
    return df


def dedup_latest_wins(
    df: pd.DataFrame,
    keys: Sequence[str],
    order_col: str = "source_gcs_path",
    quiet: bool = False,
) -> pd.DataFrame:
    """Collapse to one row per key, keeping the one from the newest file.

    MANDATORY. DoubleVerify files restate a full month-to-date window and Trade
    Desk files restate the full year, so any multi-day window contains the same
    (date, dims) row many times over. BigQuery MERGE errors out when the source
    has duplicate key rows, so this must happen before records are returned.

    order_col defaults to source_gcs_path, whose embedded <YYYY-MM-DD> folder
    sorts lexicographically in chronological order -- the same "latest delivered
    file wins" rule the origin script applied via QUALIFY ROW_NUMBER().

    quiet suppresses the per-call summary line. The extractor folds in batches
    (see GcsReportExtractor.extract_table), which would otherwise emit one line
    per batch and bury the run log; it logs its own single total instead.
    """
    if df.empty:
        return df

    usable = [k for k in keys if k in df.columns]
    if not usable:
        logger.warning("No usable dedup keys in frame; skipping dedup")
        return df
    if len(usable) != len(keys):
        missing = [k for k in keys if k not in df.columns]
        logger.warning("Dedup keys missing from frame (ignored): %s", missing)

    before = len(df)
    if order_col in df.columns:
        df = df.sort_values(order_col, kind="mergesort")
    # keep="last" == newest source file wins.
    df = df.drop_duplicates(subset=usable, keep="last").reset_index(drop=True)

    if before != len(df) and not quiet:
        logger.info(
            "Dedup: %s -> %s rows (%.1f%% collapsed, latest file wins)",
            f"{before:,}",
            f"{len(df):,}",
            (1 - len(df) / before) * 100,
        )
    return df


def frame_to_records(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """DataFrame -> list of dicts with pandas NA sentinels turned into None."""
    if df.empty:
        return []
    return df.astype(object).where(pd.notna(df), None).to_dict(orient="records")


def classify_html_capture(
    blob_name: str,
    html_expected_before: Optional[str],
) -> Tuple[str, str]:
    """Decide how to report an HTML file found where a report was expected.

    Returns (severity, message). Captures inside the known-bad backlog window
    are routine; anything at or after the cutover date means the email capture
    pipeline has broken AGAIN and should escalate.
    """
    folder = day_folder_of(blob_name) or ""
    if html_expected_before and folder and folder < str(html_expected_before):
        return "known", (
            f"{blob_name}: HTML login page (known broken feed before "
            f"{html_expected_before})"
        )
    return "regression", (
        f"{blob_name}: HTML login page captured on {folder or 'unknown date'} -- "
        f"the email ingestion is capturing login pages instead of report data"
    )


def read_header(payload: bytes, reader: str) -> List[str]:
    """Return just the snake_cased header row (used by diagnostics/tests)."""
    if reader == "csv":
        text = payload.decode("utf-8-sig", errors="replace")
        try:
            return [snake(c) for c in next(csv.reader(io.StringIO(text)))]
        except StopIteration:
            return []
    df = read_source_file(payload, reader)
    return [snake(str(c)) for c in (df.columns if df is not None else [])]


__all__ = [
    "LINEAGE_COLUMNS",
    "ResourceSpec",
    "assert_required_columns",
    "build_resource_specs",
    "classify_html_capture",
    "coalesce_key_columns",
    "day_folder_of",
    "dedup_latest_wins",
    "frame_to_records",
    "is_html",
    "list_day_folder_blobs",
    "normalize_frame",
    "read_header",
    "read_source_file",
    "route_blob",
    "snake",
    "stamp_lineage",
]
