#!/usr/bin/env python3
"""
BioNews Data Landscape Excel Generator

Creates an Excel file showing all identity nodes in the Bionews identity graph,
their neighbors, entity types, domains, and source tables/columns.

Output: shared/bionews_data_landscape.xlsx
"""

import os
import sys
from datetime import datetime

# Attempt to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Define identity sources and their metadata
IDENTITY_SOURCES = [
    # Priority 1 sources (highest value identifiers)
    {
        "mapping": "bionews_id <-> aim_dgid",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched",
        "neighbor_type": "hcp",
        "neighbor_domain": "iqvia",
        "neighbor_column": "aim_dgid",
        "neighbor_table": "page_load_sessions",
        "priority": 1,
        "section": "Priority 1 - Healthcare Identity"
    },
    {
        "mapping": "bionews_id <-> client_id",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched",
        "neighbor_type": "browser",
        "neighbor_domain": "google_analytics",
        "neighbor_column": "client_id",
        "neighbor_table": "page_load_sessions",
        "priority": 2,
        "section": "Priority 1 - Healthcare Identity"
    },
    {
        "mapping": "bionews_id <-> clarity_user",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched",
        "neighbor_type": "browser",
        "neighbor_domain": "microsoft",
        "neighbor_column": "clarity_user",
        "neighbor_table": "page_load_sessions",
        "priority": 3,
        "section": "Priority 2 - Cross-Platform Identity"
    },
    {
        "mapping": "bionews_id <-> fbp",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched",
        "neighbor_type": "browser",
        "neighbor_domain": "meta",
        "neighbor_column": "fbp",
        "neighbor_table": "page_load_sessions",
        "priority": 4,
        "section": "Priority 2 - Cross-Platform Identity"
    },
    {
        "mapping": "bionews_id <-> fbc",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "click",
        "neighbor_domain": "meta",
        "neighbor_column": "_fbc (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 5,
        "section": "Priority 2 - Cross-Platform Identity"
    },
    {
        "mapping": "bionews_id <-> fbclid",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched",
        "neighbor_type": "click",
        "neighbor_domain": "meta",
        "neighbor_column": "fbclid",
        "neighbor_table": "page_load_sessions",
        "priority": 6,
        "section": "Priority 2 - Cross-Platform Identity"
    },
    {
        "mapping": "bionews_id <-> dmd_tag",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "marketing",
        "neighbor_domain": "dmd",
        "neighbor_column": "dmd-tag (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 7,
        "section": "Priority 3 - Marketing Identity"
    },
    {
        "mapping": "bionews_id <-> gcl_au",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "conversion",
        "neighbor_domain": "google_ads",
        "neighbor_column": "_gcl_au (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 8,
        "section": "Priority 3 - Marketing Identity"
    },
    {
        "mapping": "bionews_id <-> gcl_aw",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "click",
        "neighbor_domain": "google_ads",
        "neighbor_column": "_gcl_aw (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 9,
        "section": "Priority 3 - Marketing Identity"
    },
    {
        "mapping": "bionews_id <-> kla_id",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "email_user",
        "neighbor_domain": "klaviyo",
        "neighbor_column": "__kla_id (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 10,
        "section": "Priority 3 - Marketing Identity"
    },
    {
        "mapping": "bionews_id <-> gpi",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "publisher",
        "neighbor_domain": "google",
        "neighbor_column": "__gpi (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 11,
        "section": "Priority 4 - Analytics Identity"
    },
    {
        "mapping": "bionews_id <-> ga",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "browser",
        "neighbor_domain": "google_analytics",
        "neighbor_column": "_ga (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 12,
        "section": "Priority 4 - Analytics Identity"
    },
    {
        "mapping": "bionews_id <-> chartbeat",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "engagement",
        "neighbor_domain": "chartbeat",
        "neighbor_column": "_cb (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 13,
        "section": "Priority 4 - Analytics Identity"
    },
    {
        "mapping": "bionews_id <-> omappvp",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "lead",
        "neighbor_domain": "optinmonster",
        "neighbor_column": "_omappvp (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 14,
        "section": "Priority 4 - Analytics Identity"
    },
    {
        "mapping": "bionews_id <-> limesurvey",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched, cookie",
        "neighbor_type": "survey",
        "neighbor_domain": "limesurvey",
        "neighbor_column": "LS_*_STATUS (cookie)",
        "neighbor_table": "page_load_sessions.all_cookies",
        "priority": 15,
        "section": "Priority 5 - Survey/Engagement"
    },
    {
        "mapping": "bionews_id <-> bing_user",
        "node_type": "person",
        "node_domain": "bionews",
        "node_column": "bionews_id",
        "node_table": "bionews_id_xref",
        "link_attr": "stitched",
        "neighbor_type": "browser",
        "neighbor_domain": "microsoft",
        "neighbor_column": "bing_user",
        "neighbor_table": "page_load_sessions",
        "priority": 16,
        "section": "Priority 5 - Survey/Engagement"
    },
    # visitor_id relationships
    {
        "mapping": "visitor_id <-> bionews_id",
        "node_type": "session",
        "node_domain": "bionews",
        "node_column": "visitor_id",
        "node_table": "page_load_sessions",
        "link_attr": "stitched",
        "neighbor_type": "person",
        "neighbor_domain": "bionews",
        "neighbor_column": "bionews_id",
        "neighbor_table": "bionews_id_xref",
        "priority": 0,
        "section": "Core Identity Hierarchy"
    },
    {
        "mapping": "visitor_id <-> session_id",
        "node_type": "session",
        "node_domain": "bionews",
        "node_column": "visitor_id",
        "node_table": "page_load_sessions",
        "link_attr": "computed",
        "neighbor_type": "session",
        "neighbor_domain": "bionews",
        "neighbor_column": "session_id",
        "neighbor_table": "page_load_sessions",
        "priority": 0,
        "section": "Core Identity Hierarchy"
    },
    {
        "mapping": "session_id <-> pvid",
        "node_type": "session",
        "node_domain": "bionews",
        "node_column": "session_id",
        "node_table": "page_load_sessions",
        "link_attr": "contains",
        "neighbor_type": "pageview",
        "neighbor_domain": "bionews",
        "neighbor_column": "pvid",
        "neighbor_table": "page_load_sessions",
        "priority": 0,
        "section": "Core Identity Hierarchy"
    },
    # Identifier co-occurrence pairs (from identifier_pairs table)
    {
        "mapping": "client_id <-> fbp",
        "node_type": "browser",
        "node_domain": "google_analytics",
        "node_column": "client_id",
        "node_table": "identifier_pairs",
        "link_attr": "co_occurrence",
        "neighbor_type": "browser",
        "neighbor_domain": "meta",
        "neighbor_column": "fbp",
        "neighbor_table": "identifier_pairs",
        "priority": 0,
        "section": "Identifier Co-Occurrences"
    },
    {
        "mapping": "client_id <-> clarity_user",
        "node_type": "browser",
        "node_domain": "google_analytics",
        "node_column": "client_id",
        "node_table": "identifier_pairs",
        "link_attr": "co_occurrence",
        "neighbor_type": "browser",
        "neighbor_domain": "microsoft",
        "neighbor_column": "clarity_user",
        "neighbor_table": "identifier_pairs",
        "priority": 0,
        "section": "Identifier Co-Occurrences"
    },
    {
        "mapping": "fbp <-> clarity_user",
        "node_type": "browser",
        "node_domain": "meta",
        "node_column": "fbp",
        "node_table": "identifier_pairs",
        "link_attr": "co_occurrence",
        "neighbor_type": "browser",
        "neighbor_domain": "microsoft",
        "neighbor_column": "clarity_user",
        "neighbor_table": "identifier_pairs",
        "priority": 0,
        "section": "Identifier Co-Occurrences"
    },
    {
        "mapping": "client_id <-> aim_dgid",
        "node_type": "browser",
        "node_domain": "google_analytics",
        "node_column": "client_id",
        "node_table": "identifier_pairs",
        "link_attr": "co_occurrence",
        "neighbor_type": "hcp",
        "neighbor_domain": "iqvia",
        "neighbor_column": "aim_dgid",
        "neighbor_table": "identifier_pairs",
        "priority": 0,
        "section": "Identifier Co-Occurrences"
    },
]


def create_excel_landscape():
    """Create the Excel Data Landscape file."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Landscape"

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    section_font = Font(bold=True, size=10)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Headers
    headers = [
        "Client Specific Mapping",
        "Entity Type1\n(node_type)",
        "Entity Domain1\n(node_domain)",
        "Entity ID1/Column",
        "Provider/Table",
        "Entity Link Attr\n(Event Type Location)",
        "Entity Type2\n(neighbor_type)",
        "Entity Domain2\n(neighbor_domain)",
        "Entity ID2/Column",
        "Provider/Table"
    ]

    # Column widths
    col_widths = [30, 15, 20, 20, 25, 20, 15, 20, 20, 25]

    # Set column widths
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Set header row height
    ws.row_dimensions[1].height = 35

    # Group data by section
    current_section = None
    row = 2

    for source in IDENTITY_SOURCES:
        # Check if new section
        if source["section"] != current_section:
            current_section = source["section"]
            # Write section header
            cell = ws.cell(row=row, column=1, value=current_section)
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = left_align
            # Merge section header across all columns
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = section_fill
                ws.cell(row=row, column=col).border = thin_border
            row += 1

        # Write data row
        values = [
            source["mapping"],
            source["node_type"],
            source["node_domain"],
            source["node_column"],
            source["node_table"],
            source["link_attr"],
            source["neighbor_type"],
            source["neighbor_domain"],
            source["neighbor_column"],
            source["neighbor_table"]
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = left_align if col == 1 else center_align

        row += 1

    # Add a summary section
    row += 1
    summary_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    summary_font = Font(bold=True, color="FFFFFF", size=11)

    cell = ws.cell(row=row, column=1, value="SUMMARY")
    cell.fill = summary_fill
    cell.font = summary_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    for col in range(1, 11):
        ws.cell(row=row, column=col).fill = summary_fill
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # Summary stats
    summaries = [
        ("Total Identity Sources", "16"),
        ("Canonical ID (Person)", "bionews_id"),
        ("Session ID (Browser)", "visitor_id"),
        ("Page View ID", "pvid"),
        ("Session Group ID", "session_id"),
        ("Primary Source Table", "bio_acceptor_data.page_load_sessions"),
        ("Identity Graph Table", "bio_acceptor_data.bionews_id_xref"),
        ("Co-occurrence Table", "bio_acceptor_data.identifier_pairs"),
        ("Edge Table", "identity_graph.identifier_edges"),
        ("Graph Snapshot", "identity_graph.graph_snapshot_v3"),
    ]

    for label, value in summaries:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=value).font = data_font
        ws.cell(row=row, column=2).border = thin_border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    # Save
    output_path = os.path.join(os.path.dirname(__file__), "bionews_data_landscape.xlsx")
    wb.save(output_path)
    print(f"Excel Data Landscape created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_excel_landscape()
